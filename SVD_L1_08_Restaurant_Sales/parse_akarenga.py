#!/usr/bin/env python3
"""
AKARENGA パーサー v4.0 — 完璧な拠点CSV構築

拠点: AKARENGA（赤れんがテラス）
店舗: BQ（ラ・ブリック）+ RYB（ルスツ羊蹄豚）
2025年5月新設（最新拠点）

BQ ポートフォリオ:
  LUNCH / Afternoon Tea / DINNER
  その他: ルスツ羊蹄ぶた(RYB) / wolt / Uber
  営業終了後: 席料 / 食品物販 / 花束 / 売上合計

Excel構造:
  AK2025_1Q(5月): 60列（RYBセクションなし）
  AK2025_1Q(6月)以降: 88列（RYBあり）

列検出方式: Row3/Row4のキーワードで動的検出。
"""

import csv
import glob
import os
import re
import unicodedata
from datetime import date, datetime
from pathlib import Path

import openpyxl


# ============================================================
# ユーティリティ
# ============================================================

def get_numeric(ws, row, col):
    if col is None:
        return 0
    val = ws.cell(row=row, column=col).value
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return int(val)
    if isinstance(val, str):
        cleaned = val.replace(',', '').replace('¥', '').strip()
        if cleaned == '' or cleaned == '-':
            return 0
        try:
            return int(float(cleaned))
        except ValueError:
            return 0
    return 0


def get_date(ws, row, col):
    val = ws.cell(row=row, column=col).value
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, (int, float)):
        s = str(int(val))
        if len(s) == 8:
            try:
                return date(int(s[:4]), int(s[4:6]), int(s[6:8]))
            except ValueError:
                return None
    return None


def get_weekday_jp(dt):
    return ['月', '火', '水', '木', '金', '土', '日'][dt.weekday()]


# ============================================================
# 列検出エンジン — AKARENGA専用
# ============================================================

def detect_ak_columns(ws):
    """AKARENGAのRow3/Row4を走査して列位置を動的検出する。"""
    cols = {}
    max_col = ws.max_column

    # === Row3 セクション検出（Col1-60のみ＝メインデータ領域）===
    for c in range(1, min(max_col + 1, 65)):
        val = ws.cell(row=3, column=c).value
        if val is None:
            continue
        val = str(val).strip()

        if 'LUNCH' in val and 'DINNER' not in val and 'lunch_start' not in cols:
            cols['lunch_start'] = c
        elif 'Afternoon' in val and 'at_start' not in cols:
            cols['at_start'] = c
        elif 'DINNER' in val and 'dinner_start' not in cols:
            cols['dinner_start'] = c
        elif 'レストランTOTAL' in val and 'rest_total_start' not in cols:
            cols['rest_total_start'] = c
        elif 'レストラン' in val and 'ルスツ' in val:
            # 「レストラン＋ルスツ羊蹄ぶた（人数は含まず）」
            cols['combined_total_start'] = c
        elif 'ルスツ' in val and '羊蹄' in val and 'レストラン' not in val:
            # 純粋な「ルスツ羊蹄ぶた（税込）」のみ
            cols['ryb_start'] = c
        elif '営業終了後' in val:
            cols['after_close_start'] = c

    # === LUNCH内部列 ===
    ls = cols.get('lunch_start')
    if ls:
        for c in range(ls, min(ls + 10, max_col + 1)):
            val = ws.cell(row=4, column=c).value
            if val is None:
                continue
            val = str(val).strip()
            if '件数' in val and 'l_cases' not in cols:
                cols['l_cases'] = c
            elif '人数' in val and 'l_count' not in cols:
                cols['l_count'] = c
            elif '料理売上' in val and 'l_food' not in cols:
                cols['l_food'] = c
            elif '飲料売上' in val and 'l_drink' not in cols:
                cols['l_drink'] = c
            elif '合計' in val and '人数' not in val and 'l_total' not in cols:
                cols['l_total'] = c
            elif '客単価' in val and 'l_avg' not in cols:
                cols['l_avg'] = c

    # === Afternoon Tea ===
    ats = cols.get('at_start')
    if ats:
        for c in range(ats, min(ats + 10, max_col + 1)):
            val = ws.cell(row=4, column=c).value
            if val is None:
                continue
            val = str(val).strip()
            if '件数' in val and 'at_cases' not in cols:
                cols['at_cases'] = c
            elif '人数' in val and 'at_count' not in cols:
                cols['at_count'] = c
            elif '料理売上' in val or val == '料理売上':
                if 'at_food' not in cols:
                    cols['at_food'] = c
            elif '飲料売上' in val or val == '飲料売上':
                if 'at_drink' not in cols:
                    cols['at_drink'] = c
            elif '合計' in val and '人数' not in val and 'at_total' not in cols:
                cols['at_total'] = c
            elif '客単価' in val and 'at_avg' not in cols:
                cols['at_avg'] = c

    # === DINNER内部列 ===
    ds = cols.get('dinner_start')
    if ds:
        for c in range(ds, min(ds + 10, max_col + 1)):
            val = ws.cell(row=4, column=c).value
            if val is None:
                continue
            val = str(val).strip()
            if '人数' in val and 'd_count' not in cols:
                cols['d_count'] = c
            elif '料理売上' in val and 'd_food' not in cols:
                cols['d_food'] = c
            elif '飲料売上' in val and 'd_drink' not in cols:
                cols['d_drink'] = c
            elif '合計' in val and '人数' not in val and 'd_total' not in cols:
                cols['d_total'] = c
            elif '客単価' in val and 'd_avg' not in cols:
                cols['d_avg'] = c

    # === ルスツ羊蹄ぶた(RYB) ===
    rs = cols.get('ryb_start')
    if rs:
        for c in range(rs, min(rs + 10, max_col + 1)):
            val = ws.cell(row=4, column=c).value
            if val is None:
                continue
            val = str(val).strip()
            if '件数' in val and 'ryb_cases' not in cols:
                cols['ryb_cases'] = c
            elif '人数' in val and 'ryb_count' not in cols:
                cols['ryb_count'] = c
            elif '料理売上' in val or val == '料理売上':
                if 'ryb_food' not in cols:
                    cols['ryb_food'] = c
            elif '飲料売上' in val or val == '飲料売上':
                if 'ryb_drink' not in cols:
                    cols['ryb_drink'] = c
            elif '合計' in val and '人数' not in val and 'ryb_total' not in cols:
                cols['ryb_total'] = c
            elif '客単価' in val and 'ryb_avg' not in cols:
                cols['ryb_avg'] = c
            elif 'wolt' in val.lower():
                cols['wolt'] = c
                break  # woltに達したらRYBセクション終了
            elif 'uber' in val.lower():
                cols['uber'] = c
                break

    # === wolt / Uber (RYBの後) ===
    if rs and 'wolt' not in cols:
        for c in range(rs + 8, min(rs + 15, max_col + 1)):
            val = ws.cell(row=4, column=c).value
            if val is None:
                continue
            val = str(val).strip()
            if 'wolt' in val.lower() and 'wolt' not in cols:
                cols['wolt'] = c
            elif 'uber' in val.lower() and 'uber' not in cols:
                cols['uber'] = c

    # === 営業終了後 ===
    acs = cols.get('after_close_start')
    if acs:
        for c in range(acs, min(acs + 12, max_col + 1)):
            val = ws.cell(row=4, column=c).value
            if val is None:
                continue
            val = str(val).strip()
            if '売上合計' in val and '+預り金' not in val and 'grand_total' not in cols:
                cols['grand_total'] = c
            elif '客数' in val and 'ac_count' not in cols:
                cols['ac_count'] = c
            elif val == '料理' and 'ac_food' not in cols:
                cols['ac_food'] = c
            elif val == '飲料' and 'ac_drink' not in cols:
                cols['ac_drink'] = c
            elif ('席料' in val or '室料' in val) and 'seat_fee' not in cols:
                cols['seat_fee'] = c
            elif '食品物販' in val and 'goods' not in cols:
                cols['goods'] = c
            elif '花束' in val and '預り金' not in val and '招待' not in val:
                if 'flower' not in cols:
                    cols['flower'] = c
            elif '預り金' in val or '招待' in val:
                cols['deposit'] = c

    return cols


# ============================================================
# パーサー本体
# ============================================================

def parse_ak_sheet(xlsx_path, sheet_name):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name]
    cols = detect_ak_columns(ws)

    print(f"  列検出: {len(cols)}項目")
    for key in sorted(cols.keys()):
        print(f"    {key}: Col{cols[key]}")

    rows = []
    for r in range(5, ws.max_row + 1):
        dt = get_date(ws, r, 2)
        if dt is None:
            continue
        c1 = ws.cell(row=r, column=1).value
        if c1 and isinstance(c1, str) and '合計' in c1:
            continue

        row = {
            'date': dt.strftime('%Y-%m-%d'),
            'weekday': get_weekday_jp(dt),
            # BQ LUNCH
            'l_count': get_numeric(ws, r, cols.get('l_count')),
            'l_food': get_numeric(ws, r, cols.get('l_food')),
            'l_drink': get_numeric(ws, r, cols.get('l_drink')),
            'l_total': get_numeric(ws, r, cols.get('l_total')),
            'l_avg': get_numeric(ws, r, cols.get('l_avg')),
            # Afternoon Tea
            'at_count': get_numeric(ws, r, cols.get('at_count')),
            'at_food': get_numeric(ws, r, cols.get('at_food')),
            'at_drink': get_numeric(ws, r, cols.get('at_drink')),
            'at_total': get_numeric(ws, r, cols.get('at_total')),
            # BQ DINNER
            'd_count': get_numeric(ws, r, cols.get('d_count')),
            'd_food': get_numeric(ws, r, cols.get('d_food')),
            'd_drink': get_numeric(ws, r, cols.get('d_drink')),
            'd_total': get_numeric(ws, r, cols.get('d_total')),
            'd_avg': get_numeric(ws, r, cols.get('d_avg')),
            # ルスツ羊蹄ぶた
            'ryb_count': get_numeric(ws, r, cols.get('ryb_count')),
            'ryb_food': get_numeric(ws, r, cols.get('ryb_food')),
            'ryb_drink': get_numeric(ws, r, cols.get('ryb_drink')),
            'ryb_total': get_numeric(ws, r, cols.get('ryb_total')),
            # デリバリー
            'wolt': get_numeric(ws, r, cols.get('wolt')),
            'uber': get_numeric(ws, r, cols.get('uber')),
            # 営業終了後
            'seat_fee': get_numeric(ws, r, cols.get('seat_fee')),
            'goods': get_numeric(ws, r, cols.get('goods')),
            'flower': get_numeric(ws, r, cols.get('flower')),
            'grand_total': get_numeric(ws, r, cols.get('grand_total')),
        }
        rows.append(row)

    wb.close()
    return rows


# ============================================================
# メイン
# ============================================================

def main():
    base_dir = Path(__file__).parent / 'Akarenga'
    output_dir = Path(__file__).parent / 'csv_output'
    output_dir.mkdir(exist_ok=True)
    output_path = output_dir / 'AKARENGA_daily.csv'

    fieldnames = [
        'date', 'weekday',
        'l_count', 'l_food', 'l_drink', 'l_total', 'l_avg',
        'at_count', 'at_food', 'at_drink', 'at_total',
        'd_count', 'd_food', 'd_drink', 'd_total', 'd_avg',
        'ryb_count', 'ryb_food', 'ryb_drink', 'ryb_total',
        'wolt', 'uber',
        'seat_fee', 'goods', 'flower',
        'grand_total',
    ]

    all_rows = []
    seen_dates = set()

    # find xlsx files
    xlsx_files = []
    for root, dirs, files in os.walk(str(base_dir)):
        for f in files:
            if f.endswith('.xlsx') and not f.startswith('~'):
                xlsx_files.append(os.path.join(root, f))
    xlsx_files.sort()

    print(f"{'=' * 60}")
    print(f"  AKARENGA パーサー v4.0")
    print(f"  拠点: 赤れんがテラス → BQ + RYB")
    print(f"  チャネル: L/AT/D/RYB/wolt/Uber/席料/物販/花束")
    print(f"{'=' * 60}")

    for xlsx in xlsx_files:
        print(f"\n📁 {os.path.basename(xlsx)}")
        wb = openpyxl.load_workbook(xlsx, data_only=True, read_only=False)
        sheets = wb.sheetnames
        wb.close()

        for sheet in sheets:
            print(f"  📄 Sheet: {sheet}")
            try:
                rows = parse_ak_sheet(xlsx, sheet)
                new_rows = 0
                for row in rows:
                    if row['date'] not in seen_dates:
                        seen_dates.add(row['date'])
                        all_rows.append(row)
                        new_rows += 1
                print(f"  ✅ {new_rows} new days (total: {len(all_rows)})")
            except Exception as e:
                import traceback
                print(f"  ❌ Error: {e}")
                traceback.print_exc()

    all_rows.sort(key=lambda x: x['date'])

    with open(output_path, 'w', encoding='utf-8', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(all_rows)

    print(f"\n{'=' * 60}")
    print(f"  ✅ 出力: {output_path}")
    print(f"  レコード数: {len(all_rows)}")
    if all_rows:
        print(f"  期間: {all_rows[0]['date']} 〜 {all_rows[-1]['date']}")
    print(f"{'=' * 60}")

    # === 全日検証 ===
    print(f"\n{'=' * 60}")
    print(f"  全日検証 — grand_total")
    print(f"{'=' * 60}")

    errors = 0
    ok = 0
    zero = 0
    for r in all_rows:
        gt = int(r['grand_total'])
        l = int(r['l_total'])
        at = int(r['at_total'])
        d = int(r['d_total'])
        ryb = int(r['ryb_total'])
        wolt = int(r['wolt'])
        uber = int(r['uber'])
        sf = int(r['seat_fee'])
        gd = int(r['goods'])
        fl = int(r['flower'])

        # GT = L + AT + D + 席料 + 物販 + 花束（RYBとデリバリーは含まない場合もある）
        exp_bq_only = l + at + d + sf + gd + fl
        exp_all = exp_bq_only + ryb + wolt + uber

        if gt == 0 and exp_bq_only == 0:
            zero += 1
        elif gt == exp_bq_only or gt == exp_all:
            ok += 1
        else:
            if errors < 10:
                diff_bq = gt - exp_bq_only
                diff_all = gt - exp_all
                print(f"  ❌ {r['date']}: GT={gt:,} exp_bq={exp_bq_only:,}(diff={diff_bq:,}) exp_all={exp_all:,}(diff={diff_all:,})")
                print(f"     L={l:,} AT={at:,} D={d:,} RYB={ryb:,} wolt={wolt:,} uber={uber:,} 席={sf:,} 物={gd:,} 花={fl:,}")
            errors += 1

    print(f"  全日数: {len(all_rows)}")
    print(f"  休業日: {zero}")
    print(f"  ✅ 一致: {ok}")
    print(f"  ❌ 不一致: {errors}")

    # サンプル: 9/11
    print(f"\n--- サンプル検証: 9月11日 ---")
    for row in all_rows:
        if row['date'].endswith('-09-11'):
            print(f"  {row['date']} ({row['weekday']}):")
            print(f"    L={int(row['l_total']):,} AT={int(row['at_total']):,} D={int(row['d_total']):,}")
            print(f"    RYB={int(row['ryb_total']):,} wolt={int(row['wolt']):,} uber={int(row['uber']):,}")
            print(f"    席料={int(row['seat_fee']):,} 物販={int(row['goods']):,} 花束={int(row['flower']):,}")
            print(f"    grand_total={int(row['grand_total']):,}")


if __name__ == '__main__':
    main()
