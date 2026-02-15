#!/usr/bin/env python3
"""
OKURAYAMA パーサー v4.0 — 完璧な拠点CSV構築

拠点: OKURAYAMA（大倉山）
店舗:
  1. ヌーベルプース大倉山 (NP) — メイン、70列
  2. Celeste (Ce) — カフェ、10列
  3. Repos (RP) — ラウンジ、9列

NP ポートフォリオ:
  LUNCH / DINNER / その他(L) / その他(D) / 婚礼Event
  営業終了後: 席料 / 花束 / その他 / 物販 / 売上合計

Ce/RP ポートフォリオ:
  件数 / 料理 / 飲料 / 食品物販 / 合計

出力:
  OKURAYAMA_NP_daily.csv — NP単独
  OKURAYAMA_Ce_daily.csv — Ce単独
  OKURAYAMA_RP_daily.csv — RP単独
  OKURAYAMA_daily.csv    — 拠点合計（NP + Ce + RP）

列検出方式: NP=動的検出、Ce/RP=固定列（構造安定のため）
"""

import csv
import glob
import os
import re
import unicodedata
from datetime import date, datetime
from pathlib import Path

import openpyxl


# ============================================================
# ユーティリティ
# ============================================================

def get_numeric(ws, row, col):
    if col is None:
        return 0
    val = ws.cell(row=row, column=col).value
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return int(val)
    if isinstance(val, str):
        cleaned = val.replace(',', '').replace('¥', '').strip()
        if cleaned == '' or cleaned == '-':
            return 0
        try:
            return int(float(cleaned))
        except ValueError:
            return 0
    return 0


def get_date(ws, row, col):
    val = ws.cell(row=row, column=col).value
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, (int, float)):
        s = str(int(val))
        if len(s) == 8:
            try:
                return date(int(s[:4]), int(s[4:6]), int(s[6:8]))
            except ValueError:
                return None
    return None


def get_weekday_jp(dt):
    return ['月', '火', '水', '木', '金', '土', '日'][dt.weekday()]


# ============================================================
# NP列検出エンジン
# ============================================================

def detect_np_columns(ws):
    """NPのRow3/Row5を走査して列位置を動的検出する。"""
    cols = {}
    max_col = ws.max_column

    # === Row3 セクション検出 ===
    for c in range(1, max_col + 1):
        val = ws.cell(row=3, column=c).value
        if val is None:
            continue
        val = str(val).strip()
        if 'LUNCH' in val and 'lunch_start' not in cols:
            cols['lunch_start'] = c
        elif 'DINNER' in val and 'dinner_start' not in cols:
            cols['dinner_start'] = c
        elif 'レストランTOTAL' in val:
            cols['rest_total_start'] = c
        elif 'その他売上' in val:
            cols['other_start'] = c
        elif '営業終了後' in val:
            cols['after_close_start'] = c
        elif '割引' in val:
            cols['discount_start'] = c

    # === Row4 サブセクション検出（その他売上内）===
    os_start = cols.get('other_start', 0)
    if os_start:
        for c in range(os_start, min(os_start + 30, max_col + 1)):
            val = ws.cell(row=4, column=c).value
            if val is None:
                continue
            val = str(val).strip()
            if 'Lunch' in val and 'other_lunch_start' not in cols:
                cols['other_lunch_start'] = c
            elif 'Dinner' in val and 'other_dinner_start' not in cols:
                cols['other_dinner_start'] = c
            elif '婚礼' in val or 'Event' in val:
                cols['wedding_start'] = c

    # === Row5 詳細列検出 ===
    # LUNCH (セクション開始から)
    ls = cols.get('lunch_start')
    if ls:
        for c in range(ls, min(ls + 10, max_col + 1)):
            val = ws.cell(row=5, column=c).value
            if val is None:
                continue
            val = str(val).strip().split('\n')[0]
            if '件数' in val and 'l_cases' not in cols:
                cols['l_cases'] = c
            elif '人数' in val and 'l_count' not in cols:
                cols['l_count'] = c
            elif '料理売上' in val and 'l_food' not in cols:
                cols['l_food'] = c
            elif '飲料売上' in val and 'l_drink' not in cols:
                cols['l_drink'] = c
            elif '合計' in val and 'l_total' not in cols:
                cols['l_total'] = c
            elif '客単価' in val and 'l_avg' not in cols:
                cols['l_avg'] = c

    # DINNER
    ds = cols.get('dinner_start')
    if ds:
        for c in range(ds, min(ds + 10, max_col + 1)):
            val = ws.cell(row=5, column=c).value
            if val is None:
                continue
            val = str(val).strip().split('\n')[0]
            if '件数' in val and 'd_cases' not in cols:
                cols['d_cases'] = c
            elif '人数' in val and 'd_count' not in cols:
                cols['d_count'] = c
            elif '料理売上' in val and 'd_food' not in cols:
                cols['d_food'] = c
            elif '飲料売上' in val and 'd_drink' not in cols:
                cols['d_drink'] = c
            elif '合計' in val and 'd_total' not in cols:
                cols['d_total'] = c
            elif '客単価' in val and 'd_avg' not in cols:
                cols['d_avg'] = c

    # その他(Lunch) — 室料/花束/その他/物販/合計
    ols = cols.get('other_lunch_start')
    if ols:
        for c in range(ols, min(ols + 10, max_col + 1)):
            val = ws.cell(row=5, column=c).value
            if val is None:
                continue
            val = str(val).strip().split('\n')[0]
            if '室料' in val or '席料' in val:
                if 'ol_room' not in cols:
                    cols['ol_room'] = c
            elif '花束' in val and 'ol_flower' not in cols:
                cols['ol_flower'] = c
            elif 'その他' in val and 'ol_other' not in cols:
                cols['ol_other'] = c
            elif '物販' in val and 'ol_goods' not in cols:
                cols['ol_goods'] = c
            elif '合計' in val and 'ol_total' not in cols:
                cols['ol_total'] = c

    # その他(Dinner) — 室料/花束/その他/物販/合計
    ods = cols.get('other_dinner_start')
    if ods:
        for c in range(ods, min(ods + 10, max_col + 1)):
            val = ws.cell(row=5, column=c).value
            if val is None:
                continue
            val = str(val).strip().split('\n')[0]
            if '室料' in val or '席料' in val:
                if 'od_room' not in cols:
                    cols['od_room'] = c
            elif '花束' in val and 'od_flower' not in cols:
                cols['od_flower'] = c
            elif 'その他' in val and 'od_other' not in cols:
                cols['od_other'] = c
            elif '物販' in val and 'od_goods' not in cols:
                cols['od_goods'] = c
            elif '合計' in val and 'od_total' not in cols:
                cols['od_total'] = c

    # 婚礼/Event — 客数/料理/飲料/室料/花束/その他/物販/合計
    ws_start = cols.get('wedding_start')
    if ws_start:
        for c in range(ws_start, min(ws_start + 12, max_col + 1)):
            val = ws.cell(row=5, column=c).value
            if val is None:
                continue
            val = str(val).strip().split('\n')[0]
            if '客数' in val and 'w_count' not in cols:
                cols['w_count'] = c
            elif val == '料理' and 'w_food' not in cols:
                cols['w_food'] = c
            elif val == '飲料' and 'w_drink' not in cols:
                cols['w_drink'] = c
            elif '室料' in val or '席料' in val:
                if 'w_room' not in cols:
                    cols['w_room'] = c
            elif '花束' in val and 'w_flower' not in cols:
                cols['w_flower'] = c
            elif 'その他' in val and 'w_other' not in cols:
                cols['w_other'] = c
            elif '物販' in val and 'w_goods' not in cols:
                cols['w_goods'] = c
            elif '合計' in val and 'w_total' not in cols:
                cols['w_total'] = c

    # 営業終了後トータル — 客数/料理/飲料/席料/花束/その他/物販/預かり金/売上合計
    acs = cols.get('after_close_start')
    if acs:
        for c in range(acs, min(acs + 12, max_col + 1)):
            val = ws.cell(row=5, column=c).value
            if val is None:
                continue
            val = str(val).strip().split('\n')[0]
            if '売上合計' in val:
                cols['grand_total'] = c
            elif '客数' in val and 'ac_count' not in cols:
                cols['ac_count'] = c
            elif val == '料理' and 'ac_food' not in cols:
                cols['ac_food'] = c
            elif val == '飲料' and 'ac_drink' not in cols:
                cols['ac_drink'] = c
            elif '席料' in val and 'ac_seat' not in cols:
                cols['ac_seat'] = c
            elif '花束' in val and 'ac_flower' not in cols:
                cols['ac_flower'] = c
            elif 'その他' in val and 'ac_other' not in cols:
                cols['ac_other'] = c
            elif '物販' in val and 'ac_goods' not in cols:
                cols['ac_goods'] = c
            elif '預かり金' in val and 'ac_deposit' not in cols:
                cols['ac_deposit'] = c

    return cols


# ============================================================
# NP パーサー
# ============================================================

def parse_np_sheet(xlsx_path, sheet_name):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name]
    cols = detect_np_columns(ws)

    print(f"  NP列検出: {len(cols)}項目")

    rows = []
    for r in range(6, ws.max_row + 1):
        dt = get_date(ws, r, 1)
        if dt is None:
            continue
        c2 = ws.cell(row=r, column=2).value
        if c2 and isinstance(c2, str) and '合計' in c2:
            continue

        row = {
            'date': dt.strftime('%Y-%m-%d'),
            'weekday': get_weekday_jp(dt),
            # LUNCH
            'l_count': get_numeric(ws, r, cols.get('l_count')),
            'l_food': get_numeric(ws, r, cols.get('l_food')),
            'l_drink': get_numeric(ws, r, cols.get('l_drink')),
            'l_total': get_numeric(ws, r, cols.get('l_total')),
            'l_avg': get_numeric(ws, r, cols.get('l_avg')),
            # DINNER
            'd_count': get_numeric(ws, r, cols.get('d_count')),
            'd_food': get_numeric(ws, r, cols.get('d_food')),
            'd_drink': get_numeric(ws, r, cols.get('d_drink')),
            'd_total': get_numeric(ws, r, cols.get('d_total')),
            'd_avg': get_numeric(ws, r, cols.get('d_avg')),
            # その他(Lunch)
            'ol_room': get_numeric(ws, r, cols.get('ol_room')),
            'ol_flower': get_numeric(ws, r, cols.get('ol_flower')),
            'ol_other': get_numeric(ws, r, cols.get('ol_other')),
            'ol_goods': get_numeric(ws, r, cols.get('ol_goods')),
            'ol_total': get_numeric(ws, r, cols.get('ol_total')),
            # その他(Dinner)
            'od_room': get_numeric(ws, r, cols.get('od_room')),
            'od_flower': get_numeric(ws, r, cols.get('od_flower')),
            'od_other': get_numeric(ws, r, cols.get('od_other')),
            'od_goods': get_numeric(ws, r, cols.get('od_goods')),
            'od_total': get_numeric(ws, r, cols.get('od_total')),
            # 婚礼/Event
            'w_count': get_numeric(ws, r, cols.get('w_count')),
            'w_food': get_numeric(ws, r, cols.get('w_food')),
            'w_drink': get_numeric(ws, r, cols.get('w_drink')),
            'w_room': get_numeric(ws, r, cols.get('w_room')),
            'w_flower': get_numeric(ws, r, cols.get('w_flower')),
            'w_total': get_numeric(ws, r, cols.get('w_total')),
            # 営業終了後（NP売上合計）
            'seat_fee': get_numeric(ws, r, cols.get('ac_seat')),
            'flower': get_numeric(ws, r, cols.get('ac_flower')),
            'other': get_numeric(ws, r, cols.get('ac_other')),
            'goods': get_numeric(ws, r, cols.get('ac_goods')),
            'grand_total': get_numeric(ws, r, cols.get('grand_total')),
        }
        rows.append(row)

    wb.close()
    return rows


# ============================================================
# Ce/RP パーサー
# ============================================================

def parse_ce_sheet(xlsx_path, sheet_name):
    """セレステ: 10列固定"""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name]

    rows = []
    for r in range(6, ws.max_row + 1):
        dt = get_date(ws, r, 1)
        if dt is None:
            continue
        c2 = ws.cell(row=r, column=2).value
        if c2 and isinstance(c2, str) and '合計' in c2:
            continue

        row = {
            'date': dt.strftime('%Y-%m-%d'),
            'weekday': get_weekday_jp(dt),
            'count': get_numeric(ws, r, 3),
            'food': get_numeric(ws, r, 4),
            'drink': get_numeric(ws, r, 6),
            'goods': get_numeric(ws, r, 8),
            'total': get_numeric(ws, r, 9),
            'avg': get_numeric(ws, r, 10),
        }
        rows.append(row)

    wb.close()
    return rows


def parse_rp_sheet(xlsx_path, sheet_name):
    """ルポ: 9列固定"""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name]

    rows = []
    for r in range(6, ws.max_row + 1):
        dt = get_date(ws, r, 1)
        if dt is None:
            continue
        c2 = ws.cell(row=r, column=2).value
        if c2 and isinstance(c2, str) and '合計' in c2:
            continue

        row = {
            'date': dt.strftime('%Y-%m-%d'),
            'weekday': get_weekday_jp(dt),
            'count': get_numeric(ws, r, 2),
            'food': get_numeric(ws, r, 3),
            'drink': get_numeric(ws, r, 5),
            'goods': get_numeric(ws, r, 7),
            'total': get_numeric(ws, r, 8),
            'avg': get_numeric(ws, r, 9),
        }
        rows.append(row)

    wb.close()
    return rows


# ============================================================
# ファイル処理ヘルパー
# ============================================================

def find_xlsx_files(base_dir, keyword):
    """macOS HFS+ NFD正規化対応のファイル検索"""
    result = []
    keyword_nfc = unicodedata.normalize('NFC', keyword)
    for root, dirs, files in os.walk(str(base_dir)):
        for f in files:
            if f.startswith('~') or not f.endswith('.xlsx'):
                continue
            f_nfc = unicodedata.normalize('NFC', f)
            if f_nfc.startswith(keyword_nfc):
                result.append(os.path.join(root, f))
    return sorted(result)


def process_store(store_name, base_dir, parse_fn, keyword):
    """1店舗分の全Excelを処理"""
    all_rows = []
    seen_dates = set()

    xlsx_files = find_xlsx_files(base_dir, keyword)
    print(f"  ファイル数: {len(xlsx_files)}")

    for xlsx in xlsx_files:
        print(f"\n📁 {os.path.basename(xlsx)}")
        wb = openpyxl.load_workbook(xlsx, data_only=True, read_only=False)
        sheets = wb.sheetnames
        wb.close()

        for sheet in sheets:
            print(f"  📄 Sheet: {sheet}")
            try:
                rows = parse_fn(xlsx, sheet)
                new_rows = 0
                for row in rows:
                    if row['date'] not in seen_dates:
                        seen_dates.add(row['date'])
                        all_rows.append(row)
                        new_rows += 1
                print(f"  ✅ {new_rows} new days (total: {len(all_rows)})")
            except Exception as e:
                import traceback
                print(f"  ❌ Error: {e}")
                traceback.print_exc()

    all_rows.sort(key=lambda x: x['date'])
    return all_rows


# ============================================================
# メイン
# ============================================================

def main():
    base_dir = Path(__file__).parent / 'OKURAYAMA'
    output_dir = Path(__file__).parent / 'csv_output'
    output_dir.mkdir(exist_ok=True)

    print(f"{'=' * 60}")
    print(f"  OKURAYAMA パーサー v4.0")
    print(f"  拠点: 大倉山 → NP + Ce + RP")
    print(f"{'=' * 60}")

    # === NP ===
    print(f"\n{'─' * 60}")
    print(f"  [1/3] NP（ヌーベルプース）")
    print(f"{'─' * 60}")
    np_rows = process_store('NP', base_dir, parse_np_sheet, 'NP')

    np_fields = [
        'date', 'weekday',
        'l_count', 'l_food', 'l_drink', 'l_total', 'l_avg',
        'd_count', 'd_food', 'd_drink', 'd_total', 'd_avg',
        'ol_room', 'ol_flower', 'ol_other', 'ol_goods', 'ol_total',
        'od_room', 'od_flower', 'od_other', 'od_goods', 'od_total',
        'w_count', 'w_food', 'w_drink', 'w_room', 'w_flower', 'w_total',
        'seat_fee', 'flower', 'other', 'goods',
        'grand_total',
    ]
    np_path = output_dir / 'OKURAYAMA_NP_daily.csv'
    with open(np_path, 'w', encoding='utf-8', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=np_fields)
        writer.writeheader()
        writer.writerows(np_rows)
    print(f"\n  ✅ NP: {len(np_rows)}日 → {np_path}")

    # === Ce ===
    print(f"\n{'─' * 60}")
    print(f"  [2/3] Ce（セレステ）")
    print(f"{'─' * 60}")
    ce_rows = process_store('Ce', base_dir, parse_ce_sheet, 'セレステ')

    ce_fields = ['date', 'weekday', 'count', 'food', 'drink', 'goods', 'total', 'avg']
    ce_path = output_dir / 'OKURAYAMA_Ce_daily.csv'
    with open(ce_path, 'w', encoding='utf-8', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=ce_fields)
        writer.writeheader()
        writer.writerows(ce_rows)
    print(f"\n  ✅ Ce: {len(ce_rows)}日 → {ce_path}")

    # === RP ===
    print(f"\n{'─' * 60}")
    print(f"  [3/3] RP（ルポ）")
    print(f"{'─' * 60}")
    rp_rows = process_store('RP', base_dir, parse_rp_sheet, 'ルポ')

    rp_fields = ['date', 'weekday', 'count', 'food', 'drink', 'goods', 'total', 'avg']
    rp_path = output_dir / 'OKURAYAMA_RP_daily.csv'
    with open(rp_path, 'w', encoding='utf-8', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=rp_fields)
        writer.writeheader()
        writer.writerows(rp_rows)
    print(f"\n  ✅ RP: {len(rp_rows)}日 → {rp_path}")

    # === 拠点合計 ===
    print(f"\n{'─' * 60}")
    print(f"  拠点合計 (NP + Ce + RP)")
    print(f"{'─' * 60}")

    ce_dict = {r['date']: r for r in ce_rows}
    rp_dict = {r['date']: r for r in rp_rows}

    combined_fields = ['date', 'weekday', 'np_total', 'ce_total', 'rp_total', 'base_total']
    combined_rows = []

    all_dates = sorted(set(
        [r['date'] for r in np_rows] +
        [r['date'] for r in ce_rows] +
        [r['date'] for r in rp_rows]
    ))

    np_dict = {r['date']: r for r in np_rows}

    for d in all_dates:
        np = np_dict.get(d, {})
        ce = ce_dict.get(d, {})
        rp = rp_dict.get(d, {})

        np_total = int(np.get('grand_total', 0))
        ce_total = int(ce.get('total', 0))
        rp_total = int(rp.get('total', 0))

        weekday = np.get('weekday', '') or ce.get('weekday', '') or rp.get('weekday', '')

        combined_rows.append({
            'date': d,
            'weekday': weekday,
            'np_total': np_total,
            'ce_total': ce_total,
            'rp_total': rp_total,
            'base_total': np_total + ce_total + rp_total,
        })

    combined_path = output_dir / 'OKURAYAMA_daily.csv'
    with open(combined_path, 'w', encoding='utf-8', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=combined_fields)
        writer.writeheader()
        writer.writerows(combined_rows)

    print(f"  ✅ 拠点合計: {len(combined_rows)}日 → {combined_path}")

    # === 検証 ===
    print(f"\n{'=' * 60}")
    print(f"  全日検証 — NP grand_total")
    print(f"{'=' * 60}")

    errors = 0
    ok = 0
    zero = 0
    for r in np_rows:
        gt = int(r['grand_total'])
        l = int(r['l_total'])
        d = int(r['d_total'])
        sf = int(r['seat_fee'])
        fl = int(r['flower'])
        ot = int(r['other'])
        gd = int(r['goods'])
        w = int(r['w_total'])
        # GT = L + D + 席料 + 花束 + その他 + 物販（婚礼は含まない）
        exp_no_w = l + d + sf + fl + ot + gd
        exp_w_w = exp_no_w + w
        if gt == 0 and exp_no_w == 0:
            zero += 1
        elif gt == exp_no_w or gt == exp_w_w:
            ok += 1
        else:
            if errors < 15:
                print(f"  ❌ {r['date']}: GT={gt:,} exp_no_w={exp_no_w:,} exp_w_w={exp_w_w:,} diff={gt-exp_no_w:,} wedding={w:,}")
            errors += 1

    print(f"  全日数: {len(np_rows)}")
    print(f"  休業日: {zero}")
    print(f"  ✅ 一致: {ok}")
    print(f"  ❌ 不一致: {errors}")

    # サンプル: 9/11
    print(f"\n--- サンプル検証: 9月11日 ---")
    for r in combined_rows:
        if r['date'].endswith('-09-11'):
            print(f"  {r['date']} ({r['weekday']}): NP={r['np_total']:,} Ce={r['ce_total']:,} RP={r['rp_total']:,} → 拠点合計={r['base_total']:,}")


if __name__ == '__main__':
    main()
