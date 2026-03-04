#!/usr/bin/env python3
"""
TV_TOWER パーサー v4.0 — 完璧な拠点CSV構築

拠点: TV_TOWER（テレビ塔）
店舗: THE GARDEN SAPPORO HOKKAIDO GRILLE (GA)
      + BEER GARDEN (POP-UP)
      + ピコレ・ひつじにわいん (POP-UP)

ポートフォリオ:
  GA: LUNCH / DINNER / T.O・AT / 宴会
  BG/ピコレ: 件数/人数/料理/飲料/テント利用料/物販/合計
  営業終了後: 室料 / 展望台チケット / 花束

Excel構造変遷:
  TV2023_1Q: 56列（BGなし、席料/南京錠）
  TV2023_2Q〜: 64列（BG追加、室料/展望台チケット/花束登場、南京錠消滅）
  TV2024: 66-67列（物販列追加）
  TV2025: 68列（ATセクション/アフタヌーンティー実績追加）

列検出方式: Row3/Row4のキーワードで動的検出。
"""

import csv
import glob
import os
import re
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl


# ============================================================
# ユーティリティ（MOIWAYAMAと共通）
# ============================================================

def get_numeric(ws, row, col):
    """セルから数値を安全に取得"""
    if col is None:
        return 0
    val = ws.cell(row=row, column=col).value
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return int(val)
    if isinstance(val, str):
        cleaned = val.replace(',', '').replace('¥', '').strip()
        if cleaned == '' or cleaned == '-':
            return 0
        try:
            return int(float(cleaned))
        except ValueError:
            return 0
    return 0


def get_date(ws, row, col):
    """セルから日付を安全に取得"""
    val = ws.cell(row=row, column=col).value
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, (int, float)):
        s = str(int(val))
        if len(s) == 8:
            try:
                return date(int(s[:4]), int(s[4:6]), int(s[6:8]))
            except ValueError:
                return None
    return None


def get_weekday_jp(dt):
    """日本語の曜日を返す"""
    return ['月', '火', '水', '木', '金', '土', '日'][dt.weekday()]


# ============================================================
# 列検出エンジン — TV_TOWER専用
# ============================================================

def detect_tv_columns(ws):
    """
    TV_TOWERのRow3/Row4を走査してキーワードで列位置を動的検出する。
    """
    cols = {}
    max_col = ws.max_column

    # === Row3 セクション検出 ===
    for c in range(1, max_col + 1):
        val = ws.cell(row=3, column=c).value
        if val is None:
            continue
        val = str(val).strip()

        if 'LUNCH' in val and 'lunch_start' not in cols:
            cols['lunch_start'] = c
        elif 'DINNER' in val and 'dinner_start' not in cols:
            cols['dinner_start'] = c
        elif 'レストランTOTAL' in val:
            cols['rest_total_start'] = c
        elif ('EAT-IN' in val or 'アフターランチ' in val) and ('T/O' in val or 'T.O' in val):
            cols['to_start'] = c
        elif 'Wine Bar' in val or 'WINE BAR' in val or 'ワインバー' in val:
            cols['wb_start'] = c
        elif 'ビアガーデン' in val or 'ピコレ' in val:
            cols['bg_start'] = c
        elif ('L+D' in val or 'L＋D' in val) and ('TOTAL' in val):
            # 「L+D＋T/O＋宴会場+BG TOTAL」→ 宴会より先にチェック
            cols['all_total_start'] = c
        elif ('レストラン' in val) and ('T/O' in val or 'T.O' in val) and ('TOTAL' in val):
            # 「レストラン＋T/O（人数は含まず）＋宴会場TOTAL」→ 宴会より先にチェック
            cols['rest_to_total_start'] = c
        elif '宴会' in val and '宴会場' not in val:
            # 純粋な「宴会（テレビ塔ケータリング）」のみ
            cols['bq_start'] = c
        elif '営業終了後' in val:
            cols['after_close_start'] = c

    # === LUNCH内部列 ===
    ls = cols.get('lunch_start')
    if ls:
        for c in range(ls, min(ls + 10, max_col + 1)):
            val = ws.cell(row=4, column=c).value
            if val is None:
                continue
            val = str(val).strip()
            if '件数' in val and 'l_cases' not in cols:
                cols['l_cases'] = c
            elif '人数' in val and 'l_count' not in cols:
                cols['l_count'] = c
            elif '料理売上' in val and 'l_food' not in cols:
                cols['l_food'] = c
            elif '飲料売上' in val and 'l_drink' not in cols:
                cols['l_drink'] = c
            elif ('合計' in val and ('税込' in val or val == '合計')) and 'l_total' not in cols:
                if '人数' not in val:
                    cols['l_total'] = c
            elif '客単価' in val and 'l_avg' not in cols:
                cols['l_avg'] = c

    # === DINNER内部列 ===
    ds = cols.get('dinner_start')
    if ds:
        for c in range(ds, min(ds + 10, max_col + 1)):
            val = ws.cell(row=4, column=c).value
            if val is None:
                continue
            val = str(val).strip()
            if 'NVP' in val or ('件数' in val and 'd_cases' not in cols):
                cols['d_cases'] = c
            elif '人数' in val and 'd_count' not in cols:
                cols['d_count'] = c
            elif '料理売上' in val and 'd_food' not in cols:
                cols['d_food'] = c
            elif '飲料売上' in val and 'd_drink' not in cols:
                cols['d_drink'] = c
            elif ('合計' in val and ('税込' in val or val == '合計')) and 'd_total' not in cols:
                if '人数' not in val:
                    cols['d_total'] = c
            elif '客単価' in val and 'd_avg' not in cols:
                cols['d_avg'] = c

    # === T.O / アフターランチ 内部列 ===
    ts = cols.get('to_start')
    if ts:
        for c in range(ts, min(ts + 12, max_col + 1)):
            val = ws.cell(row=4, column=c).value
            if val is None:
                continue
            val = str(val).strip()
            if 'アフタヌーンティー' in val and 'at_sales' not in cols:
                cols['at_sales'] = c
            elif '合計' in val and '人数' not in val and 'to_total' not in cols:
                cols['to_total'] = c

    # === Wine Bar 内部列 ===
    wbs = cols.get('wb_start')
    if wbs:
        for c in range(wbs, min(wbs + 10, max_col + 1)):
            val = ws.cell(row=4, column=c).value
            if val is None:
                continue
            val = str(val).strip()
            if '人数' in val and 'wb_count' not in cols:
                cols['wb_count'] = c
            elif '料理売上' in val and 'wb_food' not in cols:
                cols['wb_food'] = c
            elif '飲料売上' in val and 'wb_drink' not in cols:
                cols['wb_drink'] = c
            elif '合計' in val and '人数' not in val and 'wb_total' not in cols:
                cols['wb_total'] = c
            elif '客単価' in val and 'wb_avg' not in cols:
                cols['wb_avg'] = c

    # === 宴会 内部列 ===
    bs = cols.get('bq_start')
    if bs:
        for c in range(bs, min(bs + 10, max_col + 1)):
            val = ws.cell(row=4, column=c).value
            if val is None:
                continue
            val = str(val).strip()
            if '件数' in val and 'bq_cases' not in cols:
                cols['bq_cases'] = c
            elif '人数' in val and 'bq_count' not in cols:
                cols['bq_count'] = c
            elif '料理売上' in val and 'bq_food' not in cols:
                cols['bq_food'] = c
            elif '飲料売上' in val and 'bq_drink' not in cols:
                cols['bq_drink'] = c
            elif '合計' in val and '人数' not in val and 'bq_total' not in cols:
                cols['bq_total'] = c
            elif '客単価' in val and 'bq_avg' not in cols:
                cols['bq_avg'] = c

    # === BG/ピコレ 内部列 ===
    bgs = cols.get('bg_start')
    if bgs:
        for c in range(bgs, min(bgs + 12, max_col + 1)):
            val = ws.cell(row=4, column=c).value
            if val is None:
                continue
            val = str(val).strip()
            if '件数' in val and 'bg_cases' not in cols:
                cols['bg_cases'] = c
            elif '人数' in val and 'bg_count' not in cols:
                cols['bg_count'] = c
            elif '料理売上' in val and 'bg_food' not in cols:
                cols['bg_food'] = c
            elif '料理単価' in val:
                pass  # skip
            elif '飲料売上' in val and 'bg_drink' not in cols:
                cols['bg_drink'] = c
            elif '飲料単価' in val:
                pass  # skip
            elif 'テント' in val and 'bg_tent' not in cols:
                cols['bg_tent'] = c
            elif '物販' in val and 'bg_goods' not in cols:
                cols['bg_goods'] = c
            elif '合計' in val and '人数' not in val and 'bg_total' not in cols:
                cols['bg_total'] = c
            elif '客単価' in val and 'bg_avg' not in cols:
                cols['bg_avg'] = c

    # === 「その他」列検出 ===
    at_start = cols.get('all_total_start', 0)
    if at_start:
        for c in range(at_start, min(at_start + 8, max_col + 1)):
            val = ws.cell(row=4, column=c).value
            if val and 'その他' in str(val):
                cols['other'] = c
                break

    # === 営業終了後 内部列（最重要セクション）===
    acs = cols.get('after_close_start')
    if acs:
        for c in range(acs, min(acs + 12, max_col + 1)):
            val = ws.cell(row=4, column=c).value
            if val is None:
                continue
            val = str(val).strip()

            # 売上合計を最優先でチェック
            if '売上合計' in val:
                cols['grand_total'] = c
            elif '客数' in val and 'ac_count' not in cols:
                cols['ac_count'] = c
            elif val == '料理' and 'ac_food' not in cols:
                cols['ac_food'] = c
            elif val == '飲料' and 'ac_drink' not in cols:
                cols['ac_drink'] = c
            elif '室料' in val and 'room_fee' not in cols:
                cols['room_fee'] = c
            elif '席料' in val and 'seat_fee' not in cols:
                cols['seat_fee'] = c
            elif '展望台' in val and 'ticket' not in cols:
                cols['ticket'] = c
            elif '南京錠' in val and 'lock_fee' not in cols:
                cols['lock_fee'] = c
            elif '花束' in val and '預り金' not in val and '招待' not in val and 'flower' not in cols:
                cols['flower'] = c
            elif '預り金' in val or '招待' in val:
                cols['deposit'] = c

    return cols


# ============================================================
# パーサー本体
# ============================================================

def parse_tv_sheet(xlsx_path, sheet_name):
    """TV_TOWER の1シートをパース"""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name]

    cols = detect_tv_columns(ws)

    print(f"  列検出結果:")
    for key in sorted(cols.keys()):
        print(f"    {key}: Col{cols[key]}")

    rows = []
    for r in range(5, ws.max_row + 1):
        dt = get_date(ws, r, 2)
        if dt is None:
            continue
        c1 = ws.cell(row=r, column=1).value
        if c1 and isinstance(c1, str) and '合計' in c1:
            continue

        row = {
            'date': dt.strftime('%Y-%m-%d'),
            'weekday': get_weekday_jp(dt),
            # LUNCH
            'l_count': get_numeric(ws, r, cols.get('l_count')),
            'l_food': get_numeric(ws, r, cols.get('l_food')),
            'l_drink': get_numeric(ws, r, cols.get('l_drink')),
            'l_total': get_numeric(ws, r, cols.get('l_total')),
            'l_avg': get_numeric(ws, r, cols.get('l_avg')),
            # DINNER
            'd_count': get_numeric(ws, r, cols.get('d_count')),
            'd_food': get_numeric(ws, r, cols.get('d_food')),
            'd_drink': get_numeric(ws, r, cols.get('d_drink')),
            'd_total': get_numeric(ws, r, cols.get('d_total')),
            'd_avg': get_numeric(ws, r, cols.get('d_avg')),
            # T.O / AT
            'to_total': get_numeric(ws, r, cols.get('to_total')),
            'at_sales': get_numeric(ws, r, cols.get('at_sales')),
            # Wine Bar
            'wb_count': get_numeric(ws, r, cols.get('wb_count')),
            'wb_food': get_numeric(ws, r, cols.get('wb_food')),
            'wb_drink': get_numeric(ws, r, cols.get('wb_drink')),
            # 宴会
            'bq_count': get_numeric(ws, r, cols.get('bq_count')),
            'bq_food': get_numeric(ws, r, cols.get('bq_food')),
            'bq_drink': get_numeric(ws, r, cols.get('bq_drink')),
            'bq_total': get_numeric(ws, r, cols.get('bq_total')),
            # BG/ピコレ
            'bg_count': get_numeric(ws, r, cols.get('bg_count')),
            'bg_food': get_numeric(ws, r, cols.get('bg_food')),
            'bg_drink': get_numeric(ws, r, cols.get('bg_drink')),
            'bg_tent': get_numeric(ws, r, cols.get('bg_tent')),
            'bg_goods': get_numeric(ws, r, cols.get('bg_goods')),
            'bg_total': get_numeric(ws, r, cols.get('bg_total')),
            # 営業終了後（室料/展望台/花束）
            'room_fee': get_numeric(ws, r, cols.get('room_fee')),
            'seat_fee': get_numeric(ws, r, cols.get('seat_fee')),
            'ticket': get_numeric(ws, r, cols.get('ticket')),
            'lock_fee': get_numeric(ws, r, cols.get('lock_fee')),
            'flower': get_numeric(ws, r, cols.get('flower')),
            'other': get_numeric(ws, r, cols.get('other')),
            # 売上合計（花束預り金除く）
            'grand_total': get_numeric(ws, r, cols.get('grand_total')),
        }
        rows.append(row)

    wb.close()
    return rows


# ============================================================
# メイン
# ============================================================

def main():
    base_dir = Path(__file__).parent / 'TV_TOWER'
    output_dir = Path(__file__).parent / 'csv_output'
    output_dir.mkdir(exist_ok=True)
    output_path = output_dir / 'TV_TOWER_daily.csv'

    fieldnames = [
        'date', 'weekday',
        'l_count', 'l_food', 'l_drink', 'l_total', 'l_avg',
        'd_count', 'd_food', 'd_drink', 'd_total', 'd_avg',
        'to_total', 'at_sales',
        'wb_count', 'wb_food', 'wb_drink',
        'bq_count', 'bq_food', 'bq_drink', 'bq_total',
        'bg_count', 'bg_food', 'bg_drink', 'bg_tent', 'bg_goods', 'bg_total',
        'room_fee', 'seat_fee', 'ticket', 'lock_fee', 'flower', 'other',
        'grand_total',
    ]

    all_rows = []
    seen_dates = set()

    xlsx_files = sorted(glob.glob(str(base_dir / '**/*.xlsx'), recursive=True))
    xlsx_files = [f for f in xlsx_files if not os.path.basename(f).startswith('~')]

    print(f"{'=' * 60}")
    print(f"  TV_TOWER パーサー v4.0")
    print(f"  拠点: テレビ塔 → GA + BG + ピコレ")
    print(f"  チャネル: L/D/TO/AT/WB/宴会/BG/室料/展望台/花束")
    print(f"{'=' * 60}")

    for xlsx in xlsx_files:
        print(f"\n📁 {os.path.basename(xlsx)}")
        wb = openpyxl.load_workbook(xlsx, data_only=True, read_only=False)
        sheets = wb.sheetnames
        wb.close()

        for sheet in sheets:
            print(f"  📄 Sheet: {sheet}")
            try:
                rows = parse_tv_sheet(xlsx, sheet)
                new_rows = 0
                for row in rows:
                    if row['date'] not in seen_dates:
                        seen_dates.add(row['date'])
                        all_rows.append(row)
                        new_rows += 1
                print(f"  ✅ {new_rows} new days (total: {len(all_rows)})")
            except Exception as e:
                import traceback
                print(f"  ❌ Error: {e}")
                traceback.print_exc()

    all_rows.sort(key=lambda x: x['date'])

    with open(output_path, 'w', encoding='utf-8', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(all_rows)

    print(f"\n{'=' * 60}")
    print(f"  ✅ 出力: {output_path}")
    print(f"  レコード数: {len(all_rows)}")
    if all_rows:
        print(f"  期間: {all_rows[0]['date']} 〜 {all_rows[-1]['date']}")
    print(f"{'=' * 60}")

    # サンプル検証: 9/11のデータ
    print(f"\n--- サンプル検証: 9月11日 ---")
    for row in all_rows:
        if row['date'].endswith('-09-11'):
            gt = int(row['grand_total'])
            l = int(row['l_total'])
            d = int(row['d_total'])
            to = int(row['to_total'])
            at = int(row['at_sales'])
            bq = int(row['bq_total'])
            bg = int(row['bg_total'])
            rm = int(row['room_fee'])
            sf = int(row['seat_fee'])
            tk = int(row['ticket'])
            lk = int(row['lock_fee'])
            fl = int(row['flower'])
            ot = int(row['other'])
            ch_sum = l + d + to + at + bq + bg + rm + sf + tk + lk + fl + ot
            diff = gt - ch_sum
            print(f"  {row['date']} ({row['weekday']}):")
            print(f"    L={l:>8,} D={d:>8,} TO={to:>6,} AT={at:>6,}")
            print(f"    宴会={bq:>8,} BG={bg:>8,}")
            print(f"    室料={rm:>6,} 席料={sf:>6,} 展望台={tk:>6,} 南京錠={lk:>5,} 花束={fl:>5,} 他={ot:>5,}")
            print(f"    grand_total={gt:>10,} ch_sum={ch_sum:>10,} diff={diff:,}")
            print()


if __name__ == '__main__':
    main()
