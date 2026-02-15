#!/usr/bin/env python3
"""
MOIWAYAMA パーサー v4.0 — 完璧な拠点CSV構築

拠点: MOIWAYAMA（藻岩山）
店舗: THE JEWELS (JW)
ポートフォリオ（7ch）:
  1. LUNCH
  2. DINNER
  3. TAKE OUT
  4. 席料
  5. 南京錠
  6. 花束（プラン）
  7. も〜りすカレー（食品物販）

Excel構造:
  MW2023-2024: 48列（夜景PT列なし）
  MW2025:      50列（Col4に夜景PT、Col39に食品物販追加）

列検出方式: Row3/Row4のキーワードで動的検出。固定列番号に依存しない。
"""

import csv
import glob
import os
import re
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl


# ============================================================
# ユーティリティ
# ============================================================

def get_numeric(ws, row, col):
    """セルから数値を安全に取得"""
    if col is None:
        return 0
    val = ws.cell(row=row, column=col).value
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return int(val)
    if isinstance(val, str):
        cleaned = val.replace(',', '').replace('¥', '').strip()
        if cleaned == '' or cleaned == '-':
            return 0
        try:
            return int(float(cleaned))
        except ValueError:
            return 0
    return 0


def get_date(ws, row, col):
    """セルから日付を安全に取得"""
    val = ws.cell(row=row, column=col).value
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, (int, float)):
        # 数値8桁 (20230401)
        s = str(int(val))
        if len(s) == 8:
            try:
                return date(int(s[:4]), int(s[4:6]), int(s[6:8]))
            except ValueError:
                return None
    return None


def get_weekday_jp(dt):
    """日本語の曜日を返す"""
    return ['月', '火', '水', '木', '金', '土', '日'][dt.weekday()]


# ============================================================
# 列検出エンジン
# ============================================================

def detect_columns(ws):
    """
    Row3/Row4を走査してキーワードで列位置を動的検出する。
    固定列番号に依存しない。
    """
    cols = {}
    max_col = ws.max_column

    # === Row3 セクション検出 ===
    for c in range(1, max_col + 1):
        val = ws.cell(row=3, column=c).value
        if val is None:
            continue
        val = str(val).strip()

        if 'LUNCH' in val:
            cols['lunch_start'] = c
        elif 'DINNER' in val:
            cols['dinner_start'] = c
        elif 'レストランTOTAL' in val:
            cols['rest_total_start'] = c
        elif ('レストラン＋T/O' in val or 'レストラン+T/O' in val):
            cols['rest_to_total_start'] = c
        elif ('T.O' in val or 'T/O' in val):
            # 「レストラン＋T/O」を先に処理し、純粋なT.Oセクションのみここに来る
            cols['to_start'] = c
        elif '営業終了後' in val:
            cols['after_close_start'] = c

    # === Row4 個別列検出（営業終了後セクション内）===
    after_start = cols.get('after_close_start', 0)
    if after_start:
        for c in range(after_start, min(after_start + 15, max_col + 1)):
            val = ws.cell(row=4, column=c).value
            if val is None:
                continue
            val = str(val).strip()

            if '客数' in val and 'after_close_count' not in cols:
                cols['after_close_count'] = c
            elif val == '料理' or '料理' == val.strip():
                if 'after_close_food' not in cols:
                    cols['after_close_food'] = c
            elif val == '飲料' or '飲料' == val.strip():
                if 'after_close_drink' not in cols:
                    cols['after_close_drink'] = c
            elif '席料' in val:
                cols['seat_fee'] = c
            elif '南京錠' in val:
                cols['lock_fee'] = c
            elif '売上合計' in val:
                cols['grand_total'] = c
            elif '花束' in val and '預り金' not in val and '招待' not in val:
                cols['flower'] = c
            elif '預り金' in val or '招待' in val:
                cols['deposit'] = c

    # === Row4: 食品物販（カレー）検出 ===
    # 「レストラン＋T/O TOTAL」セクション内にある
    rest_to_start = cols.get('rest_to_total_start', 0)
    if rest_to_start:
        for c in range(rest_to_start, min(rest_to_start + 10, max_col + 1)):
            val = ws.cell(row=4, column=c).value
            if val and '食品物販' in str(val):
                cols['curry'] = c
                break

    # === Row2: 席料+南京錠+花 セクション検出 ===
    for c in range(1, max_col + 1):
        val = ws.cell(row=4, column=c).value
        if val and '席料+南京錠' in str(val):
            cols['misc_total'] = c
            break

    # === LUNCH内部列 (セクション開始から相対位置) ===
    ls = cols.get('lunch_start')
    if ls:
        # Row4でLUNCHセクション内の列を特定
        for c in range(ls, min(ls + 10, max_col + 1)):
            val = ws.cell(row=4, column=c).value
            if val is None:
                continue
            val = str(val).strip()
            if '件数' in val:
                cols['l_cases'] = c
            elif '人数' in val and 'l_count' not in cols:
                cols['l_count'] = c
            elif '料理売上' in val and 'l_food' not in cols:
                cols['l_food'] = c
            elif '飲料売上' in val and 'l_drink' not in cols:
                cols['l_drink'] = c
            elif '合計' in val and '税込' in val and 'l_total' not in cols:
                cols['l_total'] = c
            elif '客単価' in val and 'l_avg' not in cols:
                cols['l_avg'] = c

    # === DINNER内部列 ===
    ds = cols.get('dinner_start')
    if ds:
        for c in range(ds, min(ds + 10, max_col + 1)):
            val = ws.cell(row=4, column=c).value
            if val is None:
                continue
            val = str(val).strip()
            if 'NVP' in val or '件数' in val:
                cols['d_cases'] = c
            elif '人数' in val and 'd_count' not in cols:
                cols['d_count'] = c
            elif '料理売上' in val and 'd_food' not in cols:
                cols['d_food'] = c
            elif '飲料売上' in val and 'd_drink' not in cols:
                cols['d_drink'] = c
            elif '合計' in val and '税込' in val and 'd_total' not in cols:
                cols['d_total'] = c
            elif '客単価' in val and 'd_avg' not in cols:
                cols['d_avg'] = c

    # === T.O内部列 ===
    ts = cols.get('to_start')
    if ts:
        for c in range(ts, min(ts + 10, max_col + 1)):
            val = ws.cell(row=4, column=c).value
            if val is None:
                continue
            val = str(val).strip()
            # T.Oの合計列: 「合計」「合計（税込）」どちらもOK
            if '合計' in val and '人数' not in val and 'to_total' not in cols:
                cols['to_total'] = c

    return cols


# ============================================================
# パーサー本体
# ============================================================

def parse_moiwayama_sheet(xlsx_path, sheet_name):
    """MOIWAYAMA の1シートをパース"""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name]

    # 動的列検出
    cols = detect_columns(ws)

    # 検出結果の表示
    print(f"  列検出結果:")
    for key in sorted(cols.keys()):
        print(f"    {key}: Col{cols[key]}")

    rows = []
    for r in range(5, ws.max_row + 1):
        # 日付取得
        dt = get_date(ws, r, 2)
        if dt is None:
            continue

        # 「合計」行をスキップ
        c1 = ws.cell(row=r, column=1).value
        if c1 and isinstance(c1, str) and '合計' in c1:
            continue

        row = {
            'date': dt.strftime('%Y-%m-%d'),
            'weekday': get_weekday_jp(dt),
            # LUNCH
            'l_count': get_numeric(ws, r, cols.get('l_count')),
            'l_food': get_numeric(ws, r, cols.get('l_food')),
            'l_drink': get_numeric(ws, r, cols.get('l_drink')),
            'l_total': get_numeric(ws, r, cols.get('l_total')),
            'l_avg': get_numeric(ws, r, cols.get('l_avg')),
            # DINNER
            'd_count': get_numeric(ws, r, cols.get('d_count')),
            'd_food': get_numeric(ws, r, cols.get('d_food')),
            'd_drink': get_numeric(ws, r, cols.get('d_drink')),
            'd_total': get_numeric(ws, r, cols.get('d_total')),
            'd_avg': get_numeric(ws, r, cols.get('d_avg')),
            # TAKE OUT
            'to_total': get_numeric(ws, r, cols.get('to_total')),
            # 席料・南京錠・花束・カレー
            'seat_fee': get_numeric(ws, r, cols.get('seat_fee')),
            'lock_fee': get_numeric(ws, r, cols.get('lock_fee')),
            'flower': get_numeric(ws, r, cols.get('flower')),
            'curry': get_numeric(ws, r, cols.get('curry')),
            # 拠点grand_total（花束預り金除く売上合計）
            'grand_total': get_numeric(ws, r, cols.get('grand_total')),
        }
        rows.append(row)

    wb.close()
    return rows


# ============================================================
# メイン
# ============================================================

def main():
    base_dir = Path(__file__).parent / 'Mt.MOIWA'
    output_dir = Path(__file__).parent / 'csv_output'
    output_dir.mkdir(exist_ok=True)
    output_path = output_dir / 'MOIWAYAMA_daily.csv'

    fieldnames = [
        'date', 'weekday',
        'l_count', 'l_food', 'l_drink', 'l_total', 'l_avg',
        'd_count', 'd_food', 'd_drink', 'd_total', 'd_avg',
        'to_total',
        'seat_fee', 'lock_fee', 'flower', 'curry',
        'grand_total',
    ]

    all_rows = []
    seen_dates = set()

    # 全Excelファイルを処理
    xlsx_files = sorted(glob.glob(str(base_dir / '**/*.xlsx'), recursive=True))
    xlsx_files = [f for f in xlsx_files if not os.path.basename(f).startswith('~')]

    print(f"{'=' * 60}")
    print(f"  MOIWAYAMA パーサー v4.0")
    print(f"  拠点: 藻岩山 → THE JEWELS (JW)")
    print(f"  ポートフォリオ: L/D/TO/席料/南京錠/花束/カレー")
    print(f"{'=' * 60}")

    for xlsx in xlsx_files:
        print(f"\n📁 {os.path.basename(xlsx)}")
        wb = openpyxl.load_workbook(xlsx, data_only=True, read_only=False)
        sheets = wb.sheetnames
        wb.close()

        for sheet in sheets:
            print(f"  📄 Sheet: {sheet}")
            try:
                rows = parse_moiwayama_sheet(xlsx, sheet)
                new_rows = 0
                for row in rows:
                    if row['date'] not in seen_dates:
                        seen_dates.add(row['date'])
                        all_rows.append(row)
                        new_rows += 1
                print(f"  ✅ {new_rows} new days (total: {len(all_rows)})")
            except Exception as e:
                print(f"  ❌ Error: {e}")

    # 日付でソート
    all_rows.sort(key=lambda x: x['date'])

    # CSV出力
    with open(output_path, 'w', encoding='utf-8', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(all_rows)

    print(f"\n{'=' * 60}")
    print(f"  ✅ 出力: {output_path}")
    print(f"  レコード数: {len(all_rows)}")
    print(f"  期間: {all_rows[0]['date']} 〜 {all_rows[-1]['date']}")
    print(f"{'=' * 60}")

    # サンプル検証: 9/11のデータ
    print(f"\n--- サンプル検証: 9月11日 ---")
    for row in all_rows:
        if row['date'].endswith('-09-11'):
            gt = int(row['grand_total'])
            ch_sum = (int(row['l_total']) + int(row['d_total']) + int(row['to_total']) +
                      int(row['seat_fee']) + int(row['lock_fee']) + int(row['flower']) +
                      int(row['curry']))
            diff = gt - ch_sum
            print(f"  {row['date']} ({row['weekday']}): "
                  f"L={int(row['l_total']):,} D={int(row['d_total']):,} "
                  f"TO={int(row['to_total']):,} 席料={int(row['seat_fee']):,} "
                  f"南京錠={int(row['lock_fee']):,} 花束={int(row['flower']):,} "
                  f"カレー={int(row['curry']):,} "
                  f"→ GT={gt:,} (ch_sum={ch_sum:,} diff={diff:,})")


if __name__ == '__main__':
    main()
