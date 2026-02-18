#!/usr/bin/env python3
"""
SVD 売上日報パーサー v3.0 — 完璧なCSV構築
全4拠点の日別売上データをExcelから読み取り、拠点別CSVを生成する。
拠点特性に基づいた動的ヘッダー検出で全パターンに対応。
"""
import openpyxl
import csv
import os
import sys
from datetime import datetime, timedelta
from pathlib import Path

BASE_DIR = Path(__file__).parent

# ============================================================
# 拠点特性定義
# ============================================================
# JW (Mt.MOIWA): ディナー特化型
#   チャネル: L / D / T.O / 席料+南京錠+花 / もーりすカレー(2025.10~)
#   列シフト: 2024.10で LUNCH が c4→c5 にシフト（夜景PT列追加）
#   　　　　  2025.10で もーりすカレー列追加
#   BGなし。宴会なし。

# GA (TV TOWER): BG爆発型
#   チャネル: L / D / T.O(→AT→WINE BAR) / 宴会 / BG / 室料 / 展望台
#   5パターン(A-E): BG有無、テント列、物販列、T/Oセクション名変化
#   花束は預り金→CSV除外

# NP (OKURAYAMA): 夏一極集中型
#   チャネル: L(+室料+花束) / D(+室料+花束) / Event(婚礼)(+室料+花束)
#   日付が数値8桁(20230401)。70列固定。

# BQ (Akarenga): 新規出店・データ蓄積中
#   チャネル: L / AT / D / RYB / 席料 / 花束
#   DINNERに件数列なし（人数のみ）
# ============================================================


def find_section_columns(ws, row_idx=3):
    """Row3のセクション名から各チャネルの開始列を動的検出"""
    sections = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=row_idx, column=col).value
        if val and isinstance(val, str):
            val_clean = val.strip().replace('\n', ' ')
            sections[col] = val_clean
    return sections


def find_column_by_keyword(sections, keyword):
    """セクション名からキーワードを含む列を検索"""
    for col, name in sections.items():
        if keyword in name:
            return col
    return None


def find_grand_total_col(ws):
    """Row4全列をスキャンして「売上合計」を含む最後の列を返す"""
    grand_total_col = None
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=4, column=c).value
        if val and isinstance(val, str) and '売上合計' in val:
            grand_total_col = c
    return grand_total_col


def get_numeric(ws, row, col):
    """セルから数値を安全に取得"""
    if col is None:
        return 0
    val = ws.cell(row=row, column=col).value
    if val is None or val == '' or isinstance(val, str):
        return 0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0


def get_date(ws, row, col, is_numeric_date=False):
    """セルから日付を安全に取得"""
    val = ws.cell(row=row, column=col).value
    if val is None:
        return None
    if is_numeric_date:
        try:
            s = str(int(val))
            return datetime(int(s[:4]), int(s[4:6]), int(s[6:8]))
        except:
            return None
    if isinstance(val, datetime):
        return val
    return None


def get_weekday_jp(dt):
    """日本語の曜日を返す"""
    return ['月', '火', '水', '木', '金', '土', '日'][dt.weekday()]


# ============================================================
# JW パーサー
# ============================================================
def parse_jw(xlsx_path, sheet_name):
    """JW (Mt.MOIWA) の1シートをパース"""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name]
    sections = find_section_columns(ws)
    
    # 動的にLUNCH列を検出
    lunch_col = find_column_by_keyword(sections, 'LUNCH')
    dinner_col = find_column_by_keyword(sections, 'DINNER')
    to_col = find_column_by_keyword(sections, 'T.O')
    total_col = find_column_by_keyword(sections, '営業終了後')
    
    if not lunch_col:
        print(f"  WARNING: LUNCH not found in {xlsx_path} / {sheet_name}")
        return []
    
    # Row4からサブ列のオフセットを確定
    # LUNCH: 件数(+0), 人数(+1), 料理(+2), 飲料(+4), 合計(+6), 客単価(+7)
    # DINNER: 人数(+1), 料理(+2), 飲料(+4), 合計(+6), 客単価(+7)
    # T.O: 個数(+0), 人数(+1), 合計(+6)
    # 営業終了後: 席料, 南京錠, 花束 は total_col から検出
    
    # 席料・南京錠・花束・もーりすカレーの列を特定
    seat_col = None
    lock_col = None
    flower_col = None
    curry_col = None
    grand_total_col = find_grand_total_col(ws)
    
    if total_col:
        for c in range(total_col, min(total_col + 12, ws.max_column + 1)):
            val = ws.cell(row=4, column=c).value
            if val and isinstance(val, str):
                if '席料' in val:
                    seat_col = c
                elif '南京錠' in val:
                    lock_col = c
                elif '花束' in val and '預り金' not in val:
                    flower_col = c
                elif 'もーりす' in val or 'カレー' in val:
                    curry_col = c
    
    rows = []
    for r in range(5, ws.max_row + 1):
        dt = get_date(ws, r, 2)
        if dt is None:
            continue
        # 合計行スキップ
        c1 = ws.cell(row=r, column=1).value
        if c1 and isinstance(c1, str) and '合計' in c1:
            continue
        
        l_total = int(get_numeric(ws, r, lunch_col + 6))
        d_total = int(get_numeric(ws, r, dinner_col + 6)) if dinner_col else 0
        to_total = int(get_numeric(ws, r, to_col + 6)) if to_col else 0
        seat_fee = int(get_numeric(ws, r, seat_col)) if seat_col else 0
        lock_fee = int(get_numeric(ws, r, lock_col)) if lock_col else 0
        flower = int(get_numeric(ws, r, flower_col)) if flower_col else 0
        # 花束は1件¥3,000〜22,000程度。¥30,000超は列検出エラー
        if flower > 30000:
            flower = 0
        curry = int(get_numeric(ws, r, curry_col)) if curry_col else 0
        
        # レストラン売上 = L + D（テイクアウト含まず）
        rest_total = l_total + d_total
        # 全体合計 = レストラン + T.O + 付帯売上
        grand_total = rest_total + to_total + seat_fee + lock_fee + flower + curry
        
        row = {
            'date': dt.strftime('%Y-%m-%d'),
            'weekday': get_weekday_jp(dt),
            'l_count': int(get_numeric(ws, r, lunch_col + 1)),  # 人数
            'l_food': int(get_numeric(ws, r, lunch_col + 2)),
            'l_drink': int(get_numeric(ws, r, lunch_col + 4)),
            'l_total': l_total,
            'l_avg': int(get_numeric(ws, r, lunch_col + 7)),
            'd_count': int(get_numeric(ws, r, dinner_col + 1)) if dinner_col else 0,
            'd_food': int(get_numeric(ws, r, dinner_col + 2)) if dinner_col else 0,
            'd_drink': int(get_numeric(ws, r, dinner_col + 4)) if dinner_col else 0,
            'd_total': d_total,
            'd_avg': int(get_numeric(ws, r, dinner_col + 7)) if dinner_col else 0,
            'to_total': to_total,
            'rest_total': rest_total,
            'seat_fee': seat_fee,
            'lock_fee': lock_fee,
            'flower': flower,
            'curry': curry,
            'grand_total': grand_total,
        }
        rows.append(row)
    
    wb.close()
    return rows


# ============================================================
# GA パーサー
# ============================================================
def parse_ga(xlsx_path, sheet_name):
    """GA (TV TOWER) の1シートをパース"""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name]
    sections = find_section_columns(ws)
    
    lunch_col = find_column_by_keyword(sections, 'LUNCH')
    dinner_col = find_column_by_keyword(sections, 'DINNER')
    to_col = find_column_by_keyword(sections, 'T/O') or find_column_by_keyword(sections, 'T.O') or find_column_by_keyword(sections, 'アフターランチ')
    banquet_col = find_column_by_keyword(sections, '宴会')
    bg_col = find_column_by_keyword(sections, 'ビアガーデン')
    total_col = find_column_by_keyword(sections, '営業終了後')
    
    if not lunch_col:
        print(f"  WARNING: LUNCH not found in {xlsx_path} / {sheet_name}")
        return []
    
    # 席料・花束預り金・売上合計の列を特定
    room_fee_col = None
    ticket_col = None
    grand_total_col = find_grand_total_col(ws)
    
    if total_col:
        for c in range(total_col, min(total_col + 12, ws.max_column + 1)):
            val = ws.cell(row=4, column=c).value
            if val and isinstance(val, str):
                if '席料' in val or '室料' in val:
                    room_fee_col = c
    
    # BG合計列を特定
    bg_total_col = None
    if bg_col:
        for c in range(bg_col, min(bg_col + 12, ws.max_column + 1)):
            val = ws.cell(row=4, column=c).value
            if val and isinstance(val, str) and '合計' in val:
                bg_total_col = c
                break
    
    rows = []
    for r in range(5, ws.max_row + 1):
        dt = get_date(ws, r, 2)
        if dt is None:
            continue
        c1 = ws.cell(row=r, column=1).value
        if c1 and isinstance(c1, str) and '合計' in c1:
            continue
        
        row = {
            'date': dt.strftime('%Y-%m-%d'),
            'weekday': get_weekday_jp(dt),
            'l_count': int(get_numeric(ws, r, lunch_col + 1)),
            'l_food': int(get_numeric(ws, r, lunch_col + 2)),
            'l_drink': int(get_numeric(ws, r, lunch_col + 4)),
            'l_total': int(get_numeric(ws, r, lunch_col + 6)),
            'l_avg': int(get_numeric(ws, r, lunch_col + 7)),
            'd_count': int(get_numeric(ws, r, dinner_col + 1)) if dinner_col else 0,
            'd_food': int(get_numeric(ws, r, dinner_col + 2)) if dinner_col else 0,
            'd_drink': int(get_numeric(ws, r, dinner_col + 4)) if dinner_col else 0,
            'd_total': int(get_numeric(ws, r, dinner_col + 6)) if dinner_col else 0,
            'd_avg': int(get_numeric(ws, r, dinner_col + 7)) if dinner_col else 0,
            'to_total': int(get_numeric(ws, r, to_col + 6)) if to_col else 0,
            'bq_count': int(get_numeric(ws, r, banquet_col + 1)) if banquet_col else 0,
            'bq_total': int(get_numeric(ws, r, banquet_col + 6)) if banquet_col else 0,
            'bg_count': int(get_numeric(ws, r, bg_col + 1)) if bg_col else 0,
            'bg_total': int(get_numeric(ws, r, bg_total_col)) if bg_total_col else 0,
            'room_fee': int(get_numeric(ws, r, room_fee_col)) if room_fee_col else 0,
            'grand_total': int(get_numeric(ws, r, grand_total_col)) if grand_total_col else 0,
        }
        rows.append(row)
    
    wb.close()
    return rows


# ============================================================
# NP パーサー
# ============================================================
def parse_np(xlsx_path, sheet_name):
    """NP (OKURAYAMA) の1シートをパース — 日付が数値8桁"""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name]
    sections = find_section_columns(ws)
    
    lunch_col = find_column_by_keyword(sections, 'LUNCH')
    dinner_col = find_column_by_keyword(sections, 'DINNER')
    other_col = find_column_by_keyword(sections, 'その他')
    total_col = find_column_by_keyword(sections, '営業終了後') or find_column_by_keyword(sections, 'トータル売上')
    
    if not lunch_col:
        print(f"  WARNING: LUNCH not found in {xlsx_path} / {sheet_name}")
        return []
    
    # Row4でL付帯/D付帯/Event列を検出
    l_sub_col = None
    d_sub_col = None
    event_col = None
    grand_total_col = None
    
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=4, column=c).value
        if val and isinstance(val, str):
            if val.strip() == 'Lunch' and c > 20:
                l_sub_col = c
            elif val.strip() == 'Dinner' and c > 20:
                d_sub_col = c
            elif '婚礼' in val or 'Event' in val:
                event_col = c
    
    grand_total_col = find_grand_total_col(ws)
    
    rows = []
    for r in range(6, ws.max_row + 1):  # NP data starts at row 6
        dt = get_date(ws, r, 1, is_numeric_date=True)
        if dt is None:
            continue
        c1 = ws.cell(row=r, column=1).value
        if c1 and isinstance(c1, str) and '合計' in c1:
            continue
        
        # NP LUNCH: 件数なし, c3=人数?, ... 動的にオフセット
        # LUNCHセクション内: 人数(+0), 料理(+1), 飲料(+3), 合計(+5), 客単価(+6)
        row = {
            'date': dt.strftime('%Y-%m-%d'),
            'weekday': get_weekday_jp(dt),
            'l_count': int(get_numeric(ws, r, lunch_col)),
            'l_food': int(get_numeric(ws, r, lunch_col + 1)),
            'l_drink': int(get_numeric(ws, r, lunch_col + 3)),
            'l_total': int(get_numeric(ws, r, lunch_col + 5)),
            'l_avg': int(get_numeric(ws, r, lunch_col + 6)),
            'd_count': int(get_numeric(ws, r, dinner_col)) if dinner_col else 0,
            'd_food': int(get_numeric(ws, r, dinner_col + 1)) if dinner_col else 0,
            'd_drink': int(get_numeric(ws, r, dinner_col + 3)) if dinner_col else 0,
            'd_total': int(get_numeric(ws, r, dinner_col + 5)) if dinner_col else 0,
            'd_avg': int(get_numeric(ws, r, dinner_col + 6)) if dinner_col else 0,
        }
        
        # L付帯: 室料, 花束
        if l_sub_col:
            row['l_room_fee'] = int(get_numeric(ws, r, l_sub_col + 1))
            row['l_flower'] = int(get_numeric(ws, r, l_sub_col + 2))
        else:
            row['l_room_fee'] = 0
            row['l_flower'] = 0
        
        # D付帯: 室料, 花束
        if d_sub_col:
            row['d_room_fee'] = int(get_numeric(ws, r, d_sub_col + 1))
            row['d_flower'] = int(get_numeric(ws, r, d_sub_col + 2))
        else:
            row['d_room_fee'] = 0
            row['d_flower'] = 0
        
        # Event
        if event_col:
            row['event_count'] = int(get_numeric(ws, r, event_col))
            row['event_food'] = int(get_numeric(ws, r, event_col + 1))
            row['event_drink'] = int(get_numeric(ws, r, event_col + 3))
            row['event_room_fee'] = int(get_numeric(ws, r, event_col + 4))
            row['event_flower'] = int(get_numeric(ws, r, event_col + 5))
            row['event_total'] = int(get_numeric(ws, r, event_col + 6))
        else:
            row['event_count'] = 0
            row['event_food'] = 0
            row['event_drink'] = 0
            row['event_room_fee'] = 0
            row['event_flower'] = 0
            row['event_total'] = 0
        
        row['grand_total'] = int(get_numeric(ws, r, grand_total_col)) if grand_total_col else 0
        rows.append(row)
    
    wb.close()
    return rows


# ============================================================
# BQ パーサー
# ============================================================
def parse_bq(xlsx_path, sheet_name):
    """BQ (Akarenga) の1シートをパース"""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name]
    sections = find_section_columns(ws)
    
    lunch_col = find_column_by_keyword(sections, 'LUNCH')
    at_col = find_column_by_keyword(sections, 'Afternoon')
    dinner_col = find_column_by_keyword(sections, 'DINNER')
    ryb_col = find_column_by_keyword(sections, 'ルスツ') or find_column_by_keyword(sections, '羊蹄')
    total_col = find_column_by_keyword(sections, '営業終了後')
    
    if not lunch_col:
        print(f"  WARNING: LUNCH not found in {xlsx_path} / {sheet_name}")
        return []
    
    # 席料・花束・売上合計の列を特定
    seat_fee_col = None
    goods_col = None
    flower_col = None
    grand_total_col = find_grand_total_col(ws)
    
    if total_col:
        for c in range(total_col, min(total_col + 12, ws.max_column + 1)):
            val = ws.cell(row=4, column=c).value
            if val and isinstance(val, str):
                if '席料' in val:
                    seat_fee_col = c
                elif '食品物販' in val:
                    goods_col = c
                elif '花束' in val and '預り金' not in val:
                    flower_col = c
    
    rows = []
    for r in range(5, ws.max_row + 1):
        dt = get_date(ws, r, 2)
        if dt is None:
            continue
        c1 = ws.cell(row=r, column=1).value
        if c1 and isinstance(c1, str) and '合計' in c1:
            continue
        
        row = {
            'date': dt.strftime('%Y-%m-%d'),
            'weekday': get_weekday_jp(dt),
            'l_count': int(get_numeric(ws, r, lunch_col + 1)),
            'l_food': int(get_numeric(ws, r, lunch_col + 2)),
            'l_drink': int(get_numeric(ws, r, lunch_col + 4)),
            'l_total': int(get_numeric(ws, r, lunch_col + 6)),
            'l_avg': int(get_numeric(ws, r, lunch_col + 7)),
            'at_count': int(get_numeric(ws, r, at_col + 1)) if at_col else 0,
            'at_food': int(get_numeric(ws, r, at_col + 2)) if at_col else 0,
            'at_drink': int(get_numeric(ws, r, at_col + 4)) if at_col else 0,
            'at_total': int(get_numeric(ws, r, at_col + 6)) if at_col else 0,
            'at_avg': int(get_numeric(ws, r, at_col + 7)) if at_col else 0,
            # DINNER: 件数列なし。人数(+0)から始まる
            'd_count': int(get_numeric(ws, r, dinner_col)),
            'd_food': int(get_numeric(ws, r, dinner_col + 1)),
            'd_drink': int(get_numeric(ws, r, dinner_col + 3)),
            'd_total': int(get_numeric(ws, r, dinner_col + 5)),
            'd_avg': int(get_numeric(ws, r, dinner_col + 6)),
            'ryb_count': int(get_numeric(ws, r, ryb_col + 1)) if ryb_col else 0,
            'ryb_food': int(get_numeric(ws, r, ryb_col + 2)) if ryb_col else 0,
            'ryb_drink': int(get_numeric(ws, r, ryb_col + 4)) if ryb_col else 0,
            'ryb_total': int(get_numeric(ws, r, ryb_col + 6)) if ryb_col else 0,
            'ryb_avg': int(get_numeric(ws, r, ryb_col + 7)) if ryb_col else 0,
            'seat_fee': int(get_numeric(ws, r, seat_fee_col)) if seat_fee_col else 0,
            'flower': int(get_numeric(ws, r, flower_col)) if flower_col else 0,
            'grand_total': int(get_numeric(ws, r, grand_total_col)) if grand_total_col else 0,
        }
        rows.append(row)
    
    wb.close()
    return rows


# ============================================================
# メイン実行
# ============================================================
def process_store(store_id, base_dir, parse_func, file_pattern):
    """1拠点の全ファイルを処理してCSVを生成"""
    all_rows = []
    xlsx_files = sorted(base_dir.glob(file_pattern))
    
    print(f"\n{'='*60}")
    print(f"  {store_id}: {len(xlsx_files)} files found")
    print(f"{'='*60}")
    
    for xlsx_file in xlsx_files:
        wb = openpyxl.load_workbook(str(xlsx_file), data_only=True)
        for sheet_name in wb.sheetnames:
            print(f"  Parsing: {xlsx_file.name} / {sheet_name}")
            try:
                rows = parse_func(str(xlsx_file), sheet_name)
                all_rows.extend(rows)
                print(f"    → {len(rows)} rows")
            except Exception as e:
                print(f"    ❌ ERROR: {e}")
        wb.close()
    
    # 日付でソート & 重複除去（同一日付はgrand_totalが大きい方を採用）
    best = {}
    for row in sorted(all_rows, key=lambda x: x['date']):
        d = row['date']
        if d not in best or row.get('grand_total', 0) > best[d].get('grand_total', 0):
            best[d] = row
    unique_rows = [best[d] for d in sorted(best.keys())]
    
    print(f"\n  Total: {len(unique_rows)} unique days")
    
    # CSV出力
    if unique_rows:
        output_path = BASE_DIR / 'csv_output' / f'{store_id}_daily.csv'
        output_path.parent.mkdir(exist_ok=True)
        fieldnames = list(unique_rows[0].keys())
        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(unique_rows)
        print(f"  ✅ Written: {output_path} ({len(fieldnames)} columns)")
    
    return unique_rows


def validate_store(store_id, rows, expected):
    """Obsidianの数字と照合"""
    print(f"\n  === {store_id} Validation ===")
    from collections import defaultdict
    monthly = defaultdict(lambda: {'grand_total': 0, 'd_count': 0, 'l_count': 0})
    yearly = defaultdict(lambda: {'grand_total': 0, 'd_count': 0})
    
    for r in rows:
        m = r['date'][:7]
        monthly[m]['grand_total'] += r['grand_total']
        monthly[m]['d_count'] += r.get('d_count', 0)
        monthly[m]['l_count'] += r.get('l_count', 0)
        
        # 年度（4月始まり）
        dt = datetime.strptime(r['date'], '%Y-%m-%d')
        fy = dt.year if dt.month >= 4 else dt.year - 1
        fy_key = f"R{fy - 2018}"  # 2023=R5, 2024=R6, 2025=R7
        yearly[fy_key]['grand_total'] += r['grand_total']
        yearly[fy_key]['d_count'] += r.get('d_count', 0)
    
    for fy in sorted(yearly.keys()):
        y = yearly[fy]
        exp = expected.get(fy, {})
        gt_ok = '✅' if exp.get('grand_total') and abs(y['grand_total'] - exp['grand_total']) < 1000 else '❓'
        d_ok = '✅' if exp.get('d_count') and y['d_count'] == exp['d_count'] else '❓'
        print(f"  {fy}: grand=¥{y['grand_total']:,.0f} {gt_ok}  D={y['d_count']}人 {d_ok}")
        if exp.get('grand_total'):
            print(f"    Expected: grand=¥{exp['grand_total']:,.0f}  D={exp.get('d_count', '?')}人")


if __name__ == '__main__':
    print("=" * 60)
    print("  SVD 売上日報パーサー v3.0")
    print("  完璧なCSV構築 — 間違いは許されない")
    print("=" * 60)
    
    # JW
    jw_rows = process_store(
        'JW', BASE_DIR / 'Mt.MOIWA', parse_jw, '**/*.xlsx'
    )
    validate_store('JW', jw_rows, {
        'R5': {'grand_total': 117123099, 'd_count': 9628},
        'R6': {'grand_total': 141090067, 'd_count': 11503},
    })
    
    # GA
    ga_rows = process_store(
        'GA', BASE_DIR / 'TV_TOWER', parse_ga, '**/*.xlsx'
    )
    validate_store('GA', ga_rows, {
        'R5': {'grand_total': 259081408, 'd_count': 5925},
        'R6': {'grand_total': 304043246, 'd_count': 6531},
    })
    
    # NP
    np_rows = process_store(
        'NP', BASE_DIR / 'OKURAYAMA', parse_np, '**/NP*.xlsx'
    )
    
    # BQ
    bq_rows = process_store(
        'BQ', BASE_DIR / 'Akarenga', parse_bq, '**/*.xlsx'
    )
    
    print("\n" + "=" * 60)
    print("  DONE")
    print("=" * 60)
