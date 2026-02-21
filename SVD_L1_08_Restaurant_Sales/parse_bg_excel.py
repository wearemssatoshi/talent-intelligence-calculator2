#!/usr/bin/env python3
"""
BG 2023 Excel Parser — 勝利の地図をSVDの資産にする
================================================================
2023_BG_Result.xlsx から以下5種のCSVを生成:
  1. BG_weather_daily.csv    — 天候データ
  2. BG_labor_daily.csv      — 労務データ
  3. BG_plan_monthly.csv     — プラン利用データ
  4. BG_reservation_monthly.csv — 予約データ
  5. BG_hourly.csv           — 時間帯別売上
"""

import openpyxl
import csv
import os
import re
from pathlib import Path

BASE_DIR = Path(__file__).parent
EXCEL_PATH = BASE_DIR / "TV_TOWER" / "TV2023" / "2023_BG_Result.xlsx"
CSV_DIR = BASE_DIR / "csv_output"

# ── Column Mappings ──
# 5-6月: extra A/B/C/S columns shift everything right by 4
LAYOUT_56 = {
    "date_col": 8,       # 日にち
    "cust": 9,
    "food": 10,
    "bev": 12,
    "tent": 14,
    "total": 16,
    "surfins": 18,
    "forking": 19,
    "timee": 20,
    "labor_total": 21,
    "labor_productivity": 22,
    "temp_low": 24,
    "temp_high": 25,
    "temp_12h": 26,
    "temp_15h": 27,
    "temp_18h": 28,
    "temp_avg": 29,
    "weather_desc": 31,
    "weather_score": 33,
}

# 7-9月: standard layout
LAYOUT_79 = {
    "date_col": 2,       # 日にち
    "cust": 5,
    "food": 6,
    "bev": 8,
    "tent": 10,
    "total": 12,
    "surfins": 14,
    "forking": 15,
    "timee": 16,
    "labor_total": 17,
    "labor_productivity": 18,
    "temp_low": 20,
    "temp_high": 21,
    "temp_12h": 22,
    "temp_15h": 23,
    "temp_18h": 24,
    "temp_avg": 25,
    "weather_desc": 27,
    "weather_score": 29,
}

MONTH_SHEETS = {
    "2023.5月実績": ("2023-05", LAYOUT_56),
    "2023.6月実績": ("2023-06", LAYOUT_56),
    "2023.7月実績": ("2023-07", LAYOUT_79),
    "2023.8月実績": ("2023-08", LAYOUT_79),
    "2023.9月実績": ("2023-09", LAYOUT_79),
}

HOURLY_SHEETS = {
    "5月時間帯別売上実績": "2023-05",
    "6月時間帯別売上実績": "2023-06",
    "7月時間帯別売上実績": "2023-07",
    "8月時間帯別売上実績": "2023-08",
    "9月時間帯別売上実績 ": "2023-09",  # note trailing space in sheet name
}


def safe_num(val, default=0):
    """Safely convert to number, handling None, errors, and non-numeric."""
    if val is None:
        return default
    if isinstance(val, str):
        if '#' in val:  # #DIV/0!, #REF! etc.
            return default
        try:
            return float(val)
        except ValueError:
            return default
    if isinstance(val, (int, float)):
        return val
    return default


def safe_str(val, default=""):
    if val is None:
        return default
    s = str(val)
    if '#' in s:
        return default
    return s.strip()


def extract_daily_data(wb):
    """Extract daily weather + labor data from monthly sheets."""
    weather_rows = []
    labor_rows = []

    for sheet_name, (month_prefix, layout) in MONTH_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            print(f"  ⚠️ Sheet not found: {sheet_name}")
            continue

        ws = wb[sheet_name]
        for r in range(1, ws.max_row + 1):
            day_val = safe_str(ws.cell(row=r, column=1).value)
            if not day_val or 'TOTAL' in day_val or 'week' in day_val or day_val == '曜日':
                continue

            date_num = safe_num(ws.cell(row=r, column=layout["date_col"]).value)
            total_sales = safe_num(ws.cell(row=r, column=layout["total"]).value)
            cust = safe_num(ws.cell(row=r, column=layout["cust"]).value)

            if date_num <= 0 or total_sales <= 0 or cust <= 0:
                continue

            date_str = f"{month_prefix}-{int(date_num):02d}"

            # Weather data
            temp_12h = safe_num(ws.cell(row=r, column=layout["temp_12h"]).value)
            temp_15h = safe_num(ws.cell(row=r, column=layout["temp_15h"]).value)
            temp_18h = safe_num(ws.cell(row=r, column=layout["temp_18h"]).value)
            temp_avg = safe_num(ws.cell(row=r, column=layout["temp_avg"]).value)
            temp_low = safe_num(ws.cell(row=r, column=layout["temp_low"]).value)
            temp_high = safe_num(ws.cell(row=r, column=layout["temp_high"]).value)
            weather_desc = safe_str(ws.cell(row=r, column=layout["weather_desc"]).value)
            weather_score = safe_num(ws.cell(row=r, column=layout["weather_score"]).value)

            weather_rows.append({
                "date": date_str,
                "weekday": day_val,
                "temp_low": round(temp_low, 1),
                "temp_high": round(temp_high, 1),
                "temp_12h": round(temp_12h, 1) if temp_12h else "",
                "temp_15h": round(temp_15h, 1) if temp_15h else "",
                "temp_18h": round(temp_18h, 1) if temp_18h else "",
                "temp_avg": round(temp_avg, 2) if temp_avg else "",
                "weather_desc": weather_desc,
                "weather_score": int(weather_score) if weather_score else "",
                "total_sales": int(total_sales),
                "customers": int(cust),
            })

            # Labor data
            surfins = safe_num(ws.cell(row=r, column=layout["surfins"]).value)
            forking = safe_num(ws.cell(row=r, column=layout["forking"]).value)
            timee = safe_num(ws.cell(row=r, column=layout["timee"]).value)
            labor_total = safe_num(ws.cell(row=r, column=layout["labor_total"]).value)
            labor_prod = safe_num(ws.cell(row=r, column=layout["labor_productivity"]).value)

            if surfins > 0 or forking > 0 or timee > 0 or labor_total > 0:
                labor_rows.append({
                    "date": date_str,
                    "weekday": day_val,
                    "surfins_hours": round(surfins, 1),
                    "forking_hours": round(forking, 1),
                    "timee_hours": round(timee, 1),
                    "total_hours": round(labor_total, 1),
                    "sales_per_hour": int(labor_prod) if labor_prod > 0 else "",
                    "total_sales": int(total_sales),
                    "customers": int(cust),
                })

    # Deduplicate by date (Excel sometimes has duplicates due to SUB TOTAL structure)
    seen_dates = set()
    weather_dedup = []
    for row in weather_rows:
        if row["date"] not in seen_dates:
            seen_dates.add(row["date"])
            weather_dedup.append(row)

    seen_dates = set()
    labor_dedup = []
    for row in labor_rows:
        if row["date"] not in seen_dates:
            seen_dates.add(row["date"])
            labor_dedup.append(row)

    return sorted(weather_dedup, key=lambda x: x["date"]), sorted(labor_dedup, key=lambda x: x["date"])


def extract_plan_data(wb):
    """Extract plan utilization data from レビュー sheet."""
    ws = wb["レビュー"]
    plan_rows = []

    # Rows 31-36: plan data by month (5月=row31, 6月=row32, ..., TOTAL=row36)
    # Cols: A=月, B=客数, C=5500プラン, D=利用率, E=6600, F=率, G=8800, H=率, I=小計
    month_map = {31: "2023-05", 32: "2023-06", 33: "2023-07", 34: "2023-08", 35: "2023-09"}

    for row_num, month_str in month_map.items():
        customers = safe_num(ws.cell(row=row_num, column=2).value)
        plan_5500 = safe_num(ws.cell(row=row_num, column=3).value)
        rate_5500 = safe_num(ws.cell(row=row_num, column=4).value)
        plan_6600 = safe_num(ws.cell(row=row_num, column=5).value)
        rate_6600 = safe_num(ws.cell(row=row_num, column=6).value)
        plan_8800 = safe_num(ws.cell(row=row_num, column=7).value)
        rate_8800 = safe_num(ws.cell(row=row_num, column=8).value)
        plan_subtotal = safe_num(ws.cell(row=row_num, column=9).value)

        if customers <= 0:
            continue

        # Get ticket and alacarte from rows 38-42
        ticket_row = row_num + 7  # 38=5月, 39=6月, 40=7月, 41=8月, 42=9月
        ticket = safe_num(ws.cell(row=ticket_row, column=3).value)
        ticket_rate = safe_num(ws.cell(row=ticket_row, column=4).value)
        plan_total = safe_num(ws.cell(row=ticket_row, column=5).value)
        plan_total_rate = safe_num(ws.cell(row=ticket_row, column=6).value)
        alacarte = safe_num(ws.cell(row=ticket_row, column=7).value)
        alacarte_rate = safe_num(ws.cell(row=ticket_row, column=8).value)

        plan_rows.append({
            "month": month_str,
            "total_customers": int(customers),
            "plan_5500": int(plan_5500),
            "rate_5500": f"{rate_5500*100:.1f}%",
            "plan_6600": int(plan_6600),
            "rate_6600": f"{rate_6600*100:.1f}%",
            "plan_8800": int(plan_8800),
            "rate_8800": f"{rate_8800*100:.1f}%",
            "ticket": int(ticket),
            "ticket_rate": f"{ticket_rate*100:.1f}%",
            "plan_total": int(plan_total),
            "plan_total_rate": f"{plan_total_rate*100:.1f}%",
            "alacarte": int(alacarte),
            "alacarte_rate": f"{alacarte_rate*100:.1f}%",
        })

    return plan_rows


def extract_reservation_data(wb):
    """Extract reservation data from レビュー sheet."""
    ws = wb["レビュー"]
    res_rows = []

    # Rows 47-51: reservation by month
    # Cols: A=月, B=組数, C=TC, D=TC利用率, E=電話, F=電話率, G=パッサージュ, H=パッサージュ率, I=Web予約率
    month_map = {47: "2023-05", 48: "2023-06", 49: "2023-07", 50: "2023-08", 51: "2023-09"}

    for row_num, month_str in month_map.items():
        total_groups = safe_num(ws.cell(row=row_num, column=2).value)
        tc = safe_num(ws.cell(row=row_num, column=3).value)
        tc_rate = safe_num(ws.cell(row=row_num, column=4).value)
        phone = safe_num(ws.cell(row=row_num, column=5).value)
        phone_rate = safe_num(ws.cell(row=row_num, column=6).value)
        passage = safe_num(ws.cell(row=row_num, column=7).value)
        passage_rate = safe_num(ws.cell(row=row_num, column=8).value)
        web_rate = safe_num(ws.cell(row=row_num, column=9).value)

        if total_groups <= 0:
            continue

        res_rows.append({
            "month": month_str,
            "total_groups": int(total_groups),
            "tc_groups": int(tc),
            "tc_rate": f"{tc_rate*100:.1f}%",
            "phone_groups": int(phone),
            "phone_rate": f"{phone_rate*100:.1f}%",
            "passage_groups": int(passage),
            "passage_rate": f"{passage_rate*100:.1f}%",
            "web_reservation_rate": f"{web_rate*100:.1f}%",
        })

    return res_rows


def extract_hourly_data(wb):
    """Extract hourly sales data from 時間帯別 sheets."""
    hourly_rows = []

    for sheet_name, month_prefix in HOURLY_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            print(f"  ⚠️ Sheet not found: {sheet_name}")
            continue

        ws = wb[sheet_name]
        for r in range(1, ws.max_row + 1):
            cell_val = safe_str(ws.cell(row=r, column=1).value)
            # Match date pattern like "2023/07/01(土)"
            match = re.match(r'(\d{4})/(\d{2})/(\d{2})\((.)\)', cell_val)
            if not match:
                continue

            date_str = f"{match.group(1)}-{match.group(2)}-{match.group(3)}"

            h12 = safe_num(ws.cell(row=r, column=2).value)
            h13 = safe_num(ws.cell(row=r, column=3).value)
            h14 = safe_num(ws.cell(row=r, column=4).value)
            h15 = safe_num(ws.cell(row=r, column=5).value)
            h16 = safe_num(ws.cell(row=r, column=6).value)
            h17 = safe_num(ws.cell(row=r, column=7).value)
            h18 = safe_num(ws.cell(row=r, column=8).value)
            h19 = safe_num(ws.cell(row=r, column=9).value)
            h20 = safe_num(ws.cell(row=r, column=10).value)
            h21 = safe_num(ws.cell(row=r, column=11).value)
            total = safe_num(ws.cell(row=r, column=12).value)

            if total > 0:
                hourly_rows.append({
                    "date": date_str,
                    "h12": int(h12),
                    "h13": int(h13),
                    "h14": int(h14),
                    "h15": int(h15),
                    "h16": int(h16),
                    "h17": int(h17),
                    "h18": int(h18),
                    "h19": int(h19),
                    "h20": int(h20),
                    "h21": int(h21),
                    "total": int(total),
                })

    return sorted(hourly_rows, key=lambda x: x["date"])


def write_csv(path, rows, fieldnames):
    with open(path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
    print(f"  ✅ {path.name}: {len(rows)} rows")


def main():
    print(f"📖 Loading {EXCEL_PATH.name}...")
    wb = openpyxl.load_workbook(str(EXCEL_PATH), data_only=True)
    print(f"  Sheets: {wb.sheetnames}")

    # ① Weather + ② Labor
    print("\n🌤️  Extracting weather & labor data...")
    weather_rows, labor_rows = extract_daily_data(wb)
    write_csv(
        CSV_DIR / "BG_weather_daily.csv",
        weather_rows,
        ["date", "weekday", "temp_low", "temp_high", "temp_12h", "temp_15h", "temp_18h",
         "temp_avg", "weather_desc", "weather_score", "total_sales", "customers"]
    )
    write_csv(
        CSV_DIR / "BG_labor_daily.csv",
        labor_rows,
        ["date", "weekday", "surfins_hours", "forking_hours", "timee_hours",
         "total_hours", "sales_per_hour", "total_sales", "customers"]
    )

    # ③ Plan
    print("\n📊 Extracting plan utilization data...")
    plan_rows = extract_plan_data(wb)
    write_csv(
        CSV_DIR / "BG_plan_monthly.csv",
        plan_rows,
        ["month", "total_customers", "plan_5500", "rate_5500", "plan_6600", "rate_6600",
         "plan_8800", "rate_8800", "ticket", "ticket_rate", "plan_total", "plan_total_rate",
         "alacarte", "alacarte_rate"]
    )

    # ④ Reservation
    print("\n📋 Extracting reservation data...")
    res_rows = extract_reservation_data(wb)
    write_csv(
        CSV_DIR / "BG_reservation_monthly.csv",
        res_rows,
        ["month", "total_groups", "tc_groups", "tc_rate", "phone_groups", "phone_rate",
         "passage_groups", "passage_rate", "web_reservation_rate"]
    )

    # ⑤ Hourly
    print("\n⏰ Extracting hourly sales data...")
    hourly_rows = extract_hourly_data(wb)
    write_csv(
        CSV_DIR / "BG_hourly.csv",
        hourly_rows,
        ["date", "h12", "h13", "h14", "h15", "h16", "h17", "h18", "h19", "h20", "h21", "total"]
    )

    # Summary
    print(f"\n{'='*60}")
    print(f"🏆 BG 2023 — 勝利の地図 CSV化完了!")
    print(f"{'='*60}")
    print(f"  ① 天候データ:     {len(weather_rows)} days")
    print(f"  ② 労務データ:     {len(labor_rows)} days")
    print(f"  ③ プラン利用:     {len(plan_rows)} months")
    print(f"  ④ 予約データ:     {len(res_rows)} months")
    print(f"  ⑤ 時間帯別売上:  {len(hourly_rows)} days")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
